import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import load_workbook

from core.live import account_report


class AccountReportExcelFormatTest(unittest.TestCase):
    def test_total_report_uses_simple_readable_excel_format(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "total.xlsx"
            frames = {
                "账户总体情况": pd.DataFrame(
                    [
                        {
                            "日期": "2026-06-11",
                            "估算权益": 10_003_595.33782,
                            "账户Gamma": -4_375_159.3513,
                            "单日盈亏/AUM": 0.00125,
                            "备注": None,
                        }
                    ]
                ),
                "持仓记录": pd.DataFrame(
                    [
                        {
                            "日期": "2026-06-11",
                            "合约名称": "科创50购6月1750",
                            "总持仓张数": 80,
                            "最新价": 0.05335,
                            "IV": 0.380617751,
                        }
                    ]
                ),
            }

            account_report._append_daily_frames_to_total_report(
                path,
                frames,
                "2026-06-11",
            )
            workbook = load_workbook(path, data_only=False)

        summary = workbook["账户总体情况"]
        positions = workbook["持仓记录"]
        self.assertEqual(summary.freeze_panes, "A2")
        self.assertEqual(positions.freeze_panes, "A2")
        self.assertEqual(summary.auto_filter.ref, f"A1:E{summary.max_row}")
        self.assertEqual(positions.auto_filter.ref, f"A1:E{positions.max_row}")
        self.assertEqual(summary.row_dimensions[1].height, 22)
        self.assertIsNone(summary["A1"].fill.fill_type)
        self.assertEqual(summary["B2"].value, 10_003_595.33782)
        self.assertEqual(summary["B2"].number_format, "#,##0.00;-#,##0.00;0.00")
        self.assertEqual(summary["C2"].number_format, "#,##0.00;-#,##0.00;0.00")
        ratio_column = next(
            cell.column_letter
            for cell in summary[1]
            if cell.value == "单日盈亏/AUM"
        )
        self.assertEqual(
            summary[f"{ratio_column}2"].number_format,
            "0.00%;-0.00%;0.00%",
        )
        self.assertEqual(positions["C2"].number_format, "#,##0;-#,##0;0")
        self.assertEqual(positions["D2"].number_format, "0.00000;-0.00000;0.00000")
        self.assertEqual(positions["E2"].number_format, "0.00%;-0.00%;0.00%")
        self.assertGreater(positions.column_dimensions["B"].width, 13)

    def test_total_report_migrates_legacy_daily_trade_sheet_name(self):
        with TemporaryDirectory() as temp_dir:
            legacy_path = Path(temp_dir) / "legacy.xlsx"
            output_path = Path(temp_dir) / "output.xlsx"
            with pd.ExcelWriter(legacy_path, engine="openpyxl") as writer:
                pd.DataFrame(
                    [{"日期": "2026-06-09", "成交编号": "old-trade"}]
                ).to_excel(writer, sheet_name="当日交易记录", index=False)

            account_report._append_daily_frames_to_total_report(
                output_path,
                {
                    "交易记录": pd.DataFrame(
                        [{"日期": "2026-06-10", "成交编号": "new-trade"}]
                    )
                },
                "2026-06-10",
                existing_path=legacy_path,
            )
            workbook = load_workbook(output_path, data_only=False)
            trades = pd.read_excel(output_path, sheet_name="交易记录")

        self.assertEqual(workbook.sheetnames, ["交易记录"])
        self.assertEqual(trades["成交编号"].tolist(), ["old-trade", "new-trade"])


if __name__ == "__main__":
    unittest.main()
