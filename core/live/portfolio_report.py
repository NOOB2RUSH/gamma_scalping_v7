from __future__ import annotations

import os
from pathlib import Path

import pandas as pd

import core
from . import account_report, market_data, portfolio_account, storage


TEMPLATE_REPORT_PATH = (
    Path(__file__).resolve().parents[2]
    / "output"
    / "live"
    / "kc50etf"
    / "20260612_154935_report.xlsx"
)
SUMMARY_TEMPLATE_SHEET = "账户总体情况"
STRATEGY_NAME_COLUMN = "策略名称"
SUMMARY_CONTRACT_CODE_COLUMN = "合约代码"
SUMMARY_AUM_COLUMN = "AUM"
STRATEGY_DISPLAY_NAMES = {
    "kc50etf": "科创50ETF华夏",
    "300etf": "沪深300ETF华泰柏瑞",
    "500etf": "中证500ETF南方",
    "50etf": "上证50ETF华夏",
}
DETAIL_SHEETS = ("持仓记录", "交易记录")
REPORT_SHEETS = (SUMMARY_TEMPLATE_SHEET, *DETAIL_SHEETS)


def _position_report_columns():
    return [
        account_report.DEFAULT_POSITION_REPORT_COLUMNS[0],
        STRATEGY_NAME_COLUMN,
        *account_report.DEFAULT_POSITION_REPORT_COLUMNS[1:],
    ]


def _report_sheet_order(products):
    return REPORT_SHEETS


def build_portfolio_report(
    account_id="default",
    products=None,
    source="snapshot",
    date=None,
    persist_history=True,
):
    products = tuple(products or core.config.available_live_products())
    payloads = {}
    errors = {}
    for product in products:
        try:
            payload = account_report.build_live_account_report(
                product,
                account_id=account_id,
                source=source,
                date=date,
                persist_history=persist_history,
            )
            payload["_portfolio_reset_date"] = _account_reset_date(product, account_id)
            payloads[product] = payload
        except Exception as exc:
            errors[product] = str(exc)
    if not payloads:
        detail = "; ".join(f"{product}: {error}" for product, error in errors.items())
        raise ValueError(f"Unable to build any live subaccount report. {detail}")
    if errors:
        detail = "; ".join(f"{product}: {error}" for product, error in errors.items())
        raise ValueError(
            "Unable to build a complete unified account report because some "
            f"products failed. {detail}"
        )

    dates = sorted({str(payload["date"]) for payload in payloads.values()})
    if len(dates) > 1:
        detail = ", ".join(
            f"{product}={payload['date']}" for product, payload in payloads.items()
        )
        raise ValueError(
            "All products must use the same valuation date for a unified account "
            f"report. {detail}"
        )
    return {
        "account_id": account_id,
        "products": list(products),
        "date": dates[-1],
        "dates": dates,
        "shared_cash": portfolio_account.shared_cash(
            account_id=account_id,
            products=products,
        ),
        "subaccounts": payloads,
        "errors": errors,
        "frames": _combined_daily_frames(payloads, errors),
    }


def write_portfolio_report(payload):
    stamp = storage.local_now_stamp()
    out_dir = storage.portfolio_output_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{stamp}_account_report.xlsx"
    existing_path = _latest_account_report_path(out_dir, before_path=path)
    if existing_path is None and TEMPLATE_REPORT_PATH.exists():
        existing_path = TEMPLATE_REPORT_PATH
    combined = _merge_with_existing(
        payload["frames"],
        existing_path,
        payloads=payload.get("subaccounts"),
        products=payload.get("products"),
    )

    temp_path = path.with_name(f"{path.stem}.{os.getpid()}.tmp{path.suffix}")
    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
        for sheet_name in _report_sheet_order(payload["products"]):
            combined[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
        _apply_template_layout(writer.book)
        account_report._format_account_report_workbook(writer.book)
    temp_path.replace(path)

    json_path = out_dir / f"{stamp}_account_daily.json"
    storage.write_json(
        json_path,
        {
            "account_id": payload["account_id"],
            "date": payload["date"],
            "dates": payload["dates"],
            "products": payload["products"],
            "shared_cash": payload.get("shared_cash"),
            "errors": payload["errors"],
            "账户总体情况": payload["frames"][SUMMARY_TEMPLATE_SHEET].to_dict(
                "records"
            ),
        },
    )
    return {"total_excel": path, "json": json_path}


def format_terminal_summary(payload):
    lines = [
        _portfolio_snapshot_time_line(payload),
        (
            f"统一账户总表 日期={payload['date']} "
            f"标的={len(payload['subaccounts'])}/{len(payload['products'])}"
        ),
        (
            f"共享现金={_fmt(payload.get('shared_cash'))} "
            "Greeks与盈亏按标的独立展示"
        ),
    ]
    summary_frame = payload["frames"].get(SUMMARY_TEMPLATE_SHEET)
    for product in payload["products"]:
        if summary_frame is None or summary_frame.empty:
            continue
        product_rows = summary_frame[
            summary_frame[STRATEGY_NAME_COLUMN].map(
                lambda value: _strategy_name_matches(value, product)
            )
        ]
        if product_rows.empty:
            continue
        row = product_rows.iloc[-1]
        lines.append(
            f"{product}: AUM={_fmt(row.get(SUMMARY_AUM_COLUMN))} "
            f"期权单日盈亏={_fmt(row.get('期权单日盈亏'))} "
            f"ETF单日盈亏={_fmt(row.get('ETF单日盈亏'))} "
            f"净单日盈亏={_fmt(row.get('净单日盈亏'))} "
            f"单日GreeksPnL={_fmt(row.get('单日GreeksPnL'))}"
        )
    for product, error in payload["errors"].items():
        lines.append(f"WARNING {product}: {error}")
    return lines


def _portfolio_snapshot_time_line(payload):
    snapshots = payload.get("subaccounts") or {}
    times = " ".join(
        (
            f"{product}="
            f"{account_report._snapshot_time_text((snapshots.get(product) or {}).get('quote_snapshot'))}"
        )
        for product in payload.get("products", [])
    )
    return f"报告快照时间: {times}" if times else "报告快照时间: 不可用"


def _combined_daily_frames(payloads, errors):
    frames_by_sheet = {}
    summary_frames = []
    position_frames = []
    trade_frames = []
    for product, payload in payloads.items():
        reset_date = _payload_reset_date(payload)
        frames = account_report._report_frames(payload)
        summary_frames.append(
            _filter_single_product_reset_rows(
                _product_summary_frame(product, payload, frames),
                reset_date,
            )
        )
        daily_frames = account_report._daily_report_frames(payload)
        for sheet_name in DETAIL_SHEETS:
            daily_frames[sheet_name] = _filter_single_product_reset_rows(
                daily_frames[sheet_name],
                reset_date,
            )
        position_frames.append(_product_position_frame(product, daily_frames["持仓记录"]))
        trade_frames.append(daily_frames["交易记录"])

    frames_by_sheet[SUMMARY_TEMPLATE_SHEET] = _concat_exact(
        summary_frames,
        _summary_report_columns(),
    )
    frames_by_sheet["持仓记录"] = _concat_exact(
        position_frames,
        _position_report_columns(),
    )
    frames_by_sheet["交易记录"] = _concat_exact(
        trade_frames,
        account_report.TRADE_COLUMNS,
    )
    frames_by_sheet[SUMMARY_TEMPLATE_SHEET] = _backfill_summary_aum_from_positions(
        frames_by_sheet[SUMMARY_TEMPLATE_SHEET],
        frames_by_sheet["持仓记录"],
    )
    return frames_by_sheet


def _product_summary_frame(product, payload, frames):
    frame = frames[SUMMARY_TEMPLATE_SHEET].copy()
    frame[STRATEGY_NAME_COLUMN] = _strategy_display_name(product)
    frame[SUMMARY_CONTRACT_CODE_COLUMN] = _strategy_contract_code(product)
    frame[SUMMARY_AUM_COLUMN] = _summary_aum_series(
        frame,
        payload,
        position_report=frames.get("持仓记录"),
    )
    if "备注" not in frame.columns:
        frame["备注"] = None
    return frame.reindex(columns=_summary_report_columns())


def _product_position_frame(product, frame):
    result = frame.copy()
    result[STRATEGY_NAME_COLUMN] = _strategy_display_name(product)
    return result.reindex(columns=_position_report_columns())


def _summary_report_columns():
    return [
        account_report.DEFAULT_SUMMARY_REPORT_COLUMNS[0],
        STRATEGY_NAME_COLUMN,
        SUMMARY_CONTRACT_CODE_COLUMN,
        *_portfolio_summary_value_columns()[1:],
        "备注",
    ]


def _portfolio_summary_value_columns():
    columns = list(account_report.DEFAULT_SUMMARY_REPORT_COLUMNS)
    try:
        columns[columns.index("估算权益")] = SUMMARY_AUM_COLUMN
    except ValueError:
        pass
    return columns


def _summary_aum_series(frame, payload, position_report=None):
    if frame.empty or "日期" not in frame.columns:
        return pd.Series(index=frame.index, dtype="float64")
    aum_by_date = {}
    aum_by_date.update(account_report._summary_aum_by_date(payload))
    if isinstance(position_report, pd.DataFrame) and not position_report.empty:
        aum_by_date.update(account_report._position_report_aum_by_date(position_report))
    dates = pd.to_datetime(frame["日期"], errors="coerce").dt.strftime("%Y-%m-%d")
    return pd.to_numeric(dates.map(aum_by_date), errors="coerce")


def _strategy_display_name(product):
    return STRATEGY_DISPLAY_NAMES.get(str(product), str(product))


def _strategy_contract_code(product):
    try:
        symbol = market_data.SSE_ETF_OPTION_SPECS[str(product)].etf_symbol
    except KeyError:
        return None
    return f"sh{symbol}"


def _canonical_strategy_name(value):
    text = str(value)
    if text in STRATEGY_DISPLAY_NAMES:
        return _strategy_display_name(text)
    reverse = {display: display for display in STRATEGY_DISPLAY_NAMES.values()}
    return reverse.get(text, text)


def _product_from_strategy_name(value):
    text = str(value)
    if text in STRATEGY_DISPLAY_NAMES:
        return text
    for product, display in STRATEGY_DISPLAY_NAMES.items():
        if text == display:
            return product
    return None


def _strategy_name_matches(value, product):
    return _canonical_strategy_name(value) == _strategy_display_name(product)


def _account_reset_date(product, account_id):
    try:
        reset_at = account_report.account_store.load_account(
            product,
            account_id=account_id,
        ).reset_at
    except Exception:
        return None
    return _normalize_reset_date(reset_at)


def _payload_reset_date(payload):
    return _normalize_reset_date(payload.get("_portfolio_reset_date"))


def _normalize_reset_date(value):
    if value is None or value == "":
        return None
    try:
        timestamp = pd.Timestamp(value)
    except (TypeError, ValueError):
        return None
    if pd.isna(timestamp):
        return None
    if timestamp.tzinfo is not None:
        timestamp = timestamp.tz_convert("Asia/Hong_Kong").tz_localize(None)
    return timestamp.normalize()


def _filter_single_product_reset_rows(frame, reset_date):
    if reset_date is None or frame.empty or "日期" not in frame.columns:
        return frame
    result = frame.copy()
    dates = pd.to_datetime(result["日期"], errors="coerce").dt.normalize()
    keep = dates.isna() | dates.ge(reset_date)
    return result.loc[keep].reset_index(drop=True)


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _concat_exact(frames, columns):
    rows = []
    for frame in frames:
        if frame is None or frame.empty:
            continue
        rows.extend(frame.reindex(columns=columns).to_dict("records"))
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame.from_records(rows, columns=columns)


def _filter_reset_history_rows(sheet_name, frame, payloads=None):
    if frame.empty or payloads is None or "日期" not in frame.columns:
        return frame
    reset_dates = {
        product: reset_date
        for product, payload in payloads.items()
        for reset_date in [_payload_reset_date(payload)]
        if reset_date is not None
    }
    if not reset_dates:
        return frame

    result = frame.copy()
    products = _reset_filter_products(sheet_name, result, payloads)
    dates = pd.to_datetime(result["日期"], errors="coerce").dt.normalize()
    keep = pd.Series(True, index=result.index)
    for product, reset_date in reset_dates.items():
        remove = products.eq(product) & dates.notna() & dates.lt(reset_date)
        keep &= ~remove
    return result.loc[keep].reset_index(drop=True)


def _reset_filter_products(sheet_name, frame, payloads):
    products = pd.Series([None] * len(frame), index=frame.index, dtype=object)
    if STRATEGY_NAME_COLUMN in frame.columns:
        by_strategy = frame[STRATEGY_NAME_COLUMN].map(_product_from_strategy_name)
        products.loc[by_strategy.notna()] = by_strategy.loc[by_strategy.notna()]

    missing = products.isna()
    if not missing.any():
        return products

    markers = _position_product_markers(payloads)
    if sheet_name == SUMMARY_TEMPLATE_SHEET:
        inferred = frame.loc[missing].apply(
            lambda row: _product_from_summary_row(row, payloads),
            axis=1,
        )
    elif sheet_name == "持仓记录":
        inferred = frame.loc[missing].apply(
            lambda row: _product_from_position_row(row, markers),
            axis=1,
        )
    elif sheet_name == "交易记录":
        inferred = frame.loc[missing].apply(
            lambda row: _product_from_trade_row(row, markers),
            axis=1,
        )
    else:
        return products

    products.loc[inferred.index] = inferred
    return products


def _merge_with_existing(current, existing_path, payloads=None, products=None):
    existing = (
        account_report._read_report_workbook(existing_path)
        if existing_path is not None and Path(existing_path).exists()
        else {}
    )
    combined = {}
    for sheet_name, current_frame in current.items():
        daily = current_frame.copy()
        columns = list(daily.columns)
        old = existing.get(sheet_name)
        if (
            sheet_name == SUMMARY_TEMPLATE_SHEET
            and old is not None
            and STRATEGY_NAME_COLUMN not in old.columns
        ):
            old = None
        if old is None and sheet_name == SUMMARY_TEMPLATE_SHEET:
            old = _legacy_product_summary_frame(
                existing,
                products or (payloads.keys() if payloads else ()),
                columns,
            )
        if old is None and sheet_name == "交易记录":
            old = existing.get("当日交易记录")
        if old is None:
            old = pd.DataFrame(columns=columns)
        old = old.reindex(columns=columns)
        if sheet_name in DETAIL_SHEETS:
            old = _restore_template_history(old, sheet_name, columns)
        old = _filter_reset_history_rows(sheet_name, old, payloads)
        daily = _filter_reset_history_rows(sheet_name, daily, payloads)
        if not daily.empty and "日期" in daily.columns:
            old = _drop_replaced_history_rows(sheet_name, old, daily)
        if old.empty:
            frame = daily
        elif daily.empty:
            frame = old
        else:
            frame = _concat_exact([old, daily], columns)
        if sheet_name == "交易记录":
            frame = _deduplicate_trade_frame(frame)
        if sheet_name == SUMMARY_TEMPLATE_SHEET:
            frame = _normalize_summary_strategy_columns(frame)
        if sheet_name == "持仓记录":
            frame = _backfill_position_strategy_name(frame, payloads)
            frame = _backfill_position_aum(frame, payloads)
            frame = _backfill_position_holding_pnl(frame, payloads)
        combined[sheet_name] = _sort_portfolio_report_frame(sheet_name, frame)
    combined[SUMMARY_TEMPLATE_SHEET] = _backfill_summary_aum_from_positions(
        combined.get(SUMMARY_TEMPLATE_SHEET),
        combined.get("持仓记录"),
    )
    return combined


def _legacy_product_summary_frame(existing, products, columns):
    if not products:
        return None
    frames = []
    for product in products:
        frame = existing.get(product)
        if frame is None or frame.empty:
            continue
        frame = frame.copy()
        frame[STRATEGY_NAME_COLUMN] = _strategy_display_name(product)
        frame[SUMMARY_CONTRACT_CODE_COLUMN] = _strategy_contract_code(product)
        frames.append(frame.reindex(columns=columns))
    if not frames:
        return None
    return _concat_exact(frames, columns)


def _normalize_summary_strategy_columns(frame):
    if frame.empty or STRATEGY_NAME_COLUMN not in frame.columns:
        return frame
    result = frame.copy()
    products = result[STRATEGY_NAME_COLUMN].map(_product_from_strategy_name)
    matched = products.notna()
    result.loc[matched, STRATEGY_NAME_COLUMN] = products.loc[matched].map(
        _strategy_display_name
    )
    if SUMMARY_CONTRACT_CODE_COLUMN in result.columns:
        missing_code = result[SUMMARY_CONTRACT_CODE_COLUMN].isna() | result[
            SUMMARY_CONTRACT_CODE_COLUMN
        ].astype(str).str.strip().eq("")
        fill_mask = matched & missing_code
        result.loc[fill_mask, SUMMARY_CONTRACT_CODE_COLUMN] = products.loc[
            fill_mask
        ].map(_strategy_contract_code)
    return result


def _backfill_position_strategy_name(frame, payloads=None):
    if (
        frame.empty
        or payloads is None
        or STRATEGY_NAME_COLUMN not in frame.columns
        or "合约代码" not in frame.columns
        or "合约名称" not in frame.columns
    ):
        return frame
    result = frame.copy()
    missing = result[STRATEGY_NAME_COLUMN].isna() | result[
        STRATEGY_NAME_COLUMN
    ].astype(str).str.strip().eq("")
    if not missing.any():
        return result
    markers = _position_product_markers(payloads)
    inferred = result.loc[missing].apply(
        lambda row: _product_from_position_row(row, markers),
        axis=1,
    )
    for index, product in inferred.dropna().items():
        result.at[index, STRATEGY_NAME_COLUMN] = _strategy_display_name(product)
    return result


def _backfill_summary_aum_from_positions(summary_frame, position_frame):
    if (
        not isinstance(summary_frame, pd.DataFrame)
        or summary_frame.empty
        or not isinstance(position_frame, pd.DataFrame)
        or position_frame.empty
    ):
        return summary_frame
    summary_required = {"日期", STRATEGY_NAME_COLUMN, SUMMARY_AUM_COLUMN}
    position_required = {
        "日期",
        STRATEGY_NAME_COLUMN,
        "总持仓张数",
        "AUM",
        "到期日",
    }
    if not summary_required.issubset(summary_frame.columns) or not position_required.issubset(
        position_frame.columns
    ):
        return summary_frame

    option_rows = position_frame.loc[position_frame["到期日"].notna()].copy()
    if option_rows.empty:
        return summary_frame
    option_rows["_summary_aum_qty"] = pd.to_numeric(
        option_rows["总持仓张数"],
        errors="coerce",
    ).fillna(0.0).abs()
    option_rows = option_rows.loc[option_rows["_summary_aum_qty"].gt(0)].copy()
    if option_rows.empty:
        return summary_frame
    option_rows["_summary_aum_date"] = pd.to_datetime(
        option_rows["日期"],
        errors="coerce",
    ).dt.strftime("%Y-%m-%d")
    option_rows["_summary_aum_strategy"] = option_rows[STRATEGY_NAME_COLUMN].map(
        _canonical_strategy_name
    )
    option_rows["_summary_aum_value"] = pd.to_numeric(
        option_rows["AUM"],
        errors="coerce",
    )
    option_rows = option_rows.dropna(
        subset=[
            "_summary_aum_date",
            "_summary_aum_strategy",
            "_summary_aum_value",
        ]
    )
    if option_rows.empty:
        return summary_frame

    group_columns = [
        "_summary_aum_date",
        "_summary_aum_strategy",
        *[
            column
            for column in ("交易方向", "到期日")
            if column in option_rows.columns
        ],
    ]
    aum_lookup = (
        option_rows.groupby(group_columns, dropna=False)["_summary_aum_value"]
        .max()
        .groupby(level=[0, 1])
        .sum()
        .to_dict()
    )
    if not aum_lookup:
        return summary_frame

    result = summary_frame.copy()
    summary_dates = pd.to_datetime(result["日期"], errors="coerce").dt.strftime(
        "%Y-%m-%d"
    )
    summary_strategies = result[STRATEGY_NAME_COLUMN].map(_canonical_strategy_name)
    values = [
        aum_lookup.get((date, strategy))
        for date, strategy in zip(summary_dates, summary_strategies)
    ]
    filled = pd.to_numeric(pd.Series(values, index=result.index), errors="coerce")
    mask = filled.notna()
    result.loc[mask, SUMMARY_AUM_COLUMN] = filled.loc[mask]
    return result


def _drop_replaced_history_rows(sheet_name, old, daily):
    if sheet_name == SUMMARY_TEMPLATE_SHEET and STRATEGY_NAME_COLUMN in daily.columns:
        if STRATEGY_NAME_COLUMN not in old.columns:
            return old
        replace_keys = _summary_replace_keys(daily)
        old_keys = _summary_replace_keys(old)
        return old.loc[~old_keys.isin(replace_keys)]
    replace_dates = set(
        pd.to_datetime(daily["日期"], errors="coerce")
        .dropna()
        .dt.strftime("%Y-%m-%d")
    )
    old_dates = pd.to_datetime(old["日期"], errors="coerce").dt.strftime("%Y-%m-%d")
    return old.loc[~old_dates.isin(replace_dates)]


def _summary_replace_keys(frame):
    dates = pd.to_datetime(frame["日期"], errors="coerce").dt.strftime("%Y-%m-%d")
    strategies = frame[STRATEGY_NAME_COLUMN].map(_canonical_strategy_name)
    return dates.fillna("").str.cat(strategies.fillna(""), sep="|")


def _sort_portfolio_report_frame(sheet_name, frame):
    result = account_report._sort_report_frame(frame)
    if (
        sheet_name != SUMMARY_TEMPLATE_SHEET
        or result.empty
        or STRATEGY_NAME_COLUMN not in result.columns
    ):
        return result
    result = result.copy()
    result["_portfolio_sort_date"] = pd.to_datetime(
        result["日期"],
        errors="coerce",
    )
    result["_portfolio_sort_strategy"] = result[STRATEGY_NAME_COLUMN].astype(str)
    result = result.sort_values(
        ["_portfolio_sort_date", "_portfolio_sort_strategy"],
        kind="stable",
    )
    return result.drop(
        columns=["_portfolio_sort_date", "_portfolio_sort_strategy"]
    ).reset_index(drop=True)


def _backfill_position_holding_pnl(frame, payloads=None):
    if frame.empty or payloads is None:
        return frame
    required = {
        "日期",
        "合约代码",
        "合约名称",
        "交易方向",
        "总持仓张数",
        "今日变化",
        "最新价",
        "持仓盈亏",
    }
    if not required.issubset(frame.columns):
        return frame
    result = frame.copy()
    product_markers = _position_product_markers(payloads)
    result["_position_date"] = pd.to_datetime(result["日期"], errors="coerce")
    result["_position_code"] = result["合约代码"].map(_position_code_key)
    result["_position_product"] = result["合约名称"].map(
        lambda name: _product_from_contract_name(name, product_markers)
    )
    result["_original_order"] = range(len(result))

    sorted_result = result.sort_values(
        ["_position_product", "_position_code", "_position_date", "_original_order"],
        kind="stable",
    )
    for _, group in sorted_result.groupby(
        ["_position_product", "_position_code"],
        dropna=False,
        sort=False,
    ):
        previous_row = None
        for index, row in group.iterrows():
            qty = _number(row.get("总持仓张数"))
            latest = _number(row.get("最新价"))
            previous_qty = _number(previous_row.get("总持仓张数")) if previous_row is not None else 0.0
            previous_latest = _number(previous_row.get("最新价")) if previous_row is not None else None
            previous_direction = (
                previous_row.get("交易方向") if previous_row is not None else row.get("交易方向")
            )
            if qty is not None:
                result.at[index, "今日变化"] = qty - (previous_qty or 0.0)
            product = row.get("_position_product")
            trade_rows = _position_trade_rows(
                payloads.get(product, {}),
                row.get("日期"),
                row.get("合约代码"),
            )
            multiplier = _position_row_multiplier(row, product_markers)
            pnl = account_report._daily_position_pnl_breakdown(
                current_qty=qty,
                current_side=_account_position_side(row.get("交易方向")),
                current_price=latest,
                previous_qty=previous_qty,
                previous_side=_account_position_side(previous_direction),
                previous_price=previous_latest,
                previous_cost=(
                    _number(previous_row.get("持仓均价"))
                    if previous_row is not None
                    else None
                ),
                trade_rows=trade_rows,
                multiplier=multiplier,
            )
            result.at[index, "持仓盈亏"] = pnl["holding_pnl"]
            previous_row = row

    return result.drop(
        columns=[
            "_position_date",
            "_position_code",
            "_position_product",
            "_original_order",
        ]
    )


def _position_trade_rows(payload, date, code):
    date_text = str(pd.Timestamp(date).date()) if not pd.isna(date) else ""
    code_key = _position_code_key(code)
    return [
        trade
        for trade in payload.get("trade_rows", [])
        if str(trade.get("日期")) == date_text
        and _position_code_key(trade.get("合约代码")) == code_key
    ]


def _account_position_side(value):
    return "short" if _position_direction_sign(value) < 0 else "long"


def _position_code_key(value):
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return text


def _position_direction_sign(direction):
    return -1.0 if str(direction or "") == "空" else 1.0


def _position_row_multiplier(row, product_markers):
    if pd.isna(row.get("到期日")):
        return 1.0
    product = _product_from_contract_name(row.get("合约名称"), product_markers)
    return _contract_multiplier(product)


def _backfill_position_aum(frame, payloads=None):
    if frame.empty or payloads is None:
        return frame
    required = {"日期", "合约名称", "总持仓张数", "到期日", "AUM"}
    if not required.issubset(frame.columns):
        return frame
    result = frame.copy()
    spot_lookup = _spot_lookup(payloads)
    product_markers = _position_product_markers(payloads)
    product_by_index = result["合约名称"].map(
        lambda name: _product_from_contract_name(name, product_markers)
    )
    option_mask = result["到期日"].notna() & product_by_index.notna()
    if not option_mask.any():
        return result

    result["_aum_product"] = product_by_index
    result["_aum_date"] = pd.to_datetime(
        result["日期"],
        errors="coerce",
    ).dt.strftime("%Y-%m-%d")
    grouping_columns = ["_aum_date", "_aum_product", "交易方向", "到期日"]
    for _, group in result.loc[option_mask].groupby(
        grouping_columns,
        dropna=False,
        sort=False,
    ):
        product = group["_aum_product"].iloc[0]
        date = group["_aum_date"].iloc[0]
        spot = spot_lookup.get((product, date))
        if spot is None:
            continue
        qty = pd.to_numeric(group["总持仓张数"], errors="coerce").abs().max()
        if pd.isna(qty):
            continue
        aum = float(qty) * _contract_multiplier(product) * float(spot)
        result.loc[group.index, "AUM"] = aum
    return result.drop(columns=["_aum_product", "_aum_date"])


def _spot_lookup(payloads):
    lookup = {}
    for product, payload in payloads.items():
        history = payload.get("summary_history")
        if not isinstance(history, pd.DataFrame) or history.empty:
            date = str(payload.get("date"))
            spot = _number(payload.get("summary", {}).get("标的价格"))
            if spot is None:
                spot = _number(payload.get("spot"))
            if spot is not None:
                lookup[(product, date)] = spot
            continue
        if "日期" not in history or "标的价格" not in history:
            continue
        for _, row in history.iterrows():
            date = pd.to_datetime(row.get("日期"), errors="coerce")
            spot = _number(row.get("标的价格"))
            if pd.isna(date) or spot is None:
                continue
            lookup[(product, str(date.date()))] = spot
    return lookup


def _position_product_markers(payloads):
    markers = {
        "50etf": ("50ETF",),
        "300etf": ("300ETF",),
        "500etf": ("500ETF",),
        "kc50etf": ("科创50",),
    }
    return {product: markers.get(product, (product,)) for product in payloads}


def _product_from_contract_name(name, markers):
    text = str(name or "")
    for product, product_markers in markers.items():
        if any(marker in text for marker in product_markers):
            return product
    return None


def _product_from_summary_row(row, payloads):
    return _product_from_security_code(
        row.get(SUMMARY_CONTRACT_CODE_COLUMN),
        payloads.keys(),
    )


def _product_from_position_row(row, markers):
    product = _product_from_contract_name(row.get("合约名称"), markers)
    if product is not None:
        return product
    return _product_from_security_code(row.get("合约代码"), markers.keys())


def _product_from_trade_row(row, markers):
    product = _product_from_contract_name(row.get("合约名称"), markers)
    if product is not None:
        return product
    return _product_from_security_code(row.get("合约代码"), markers.keys())


def _product_from_security_code(value, products):
    code = _position_code_key(value).lower()
    if code.startswith(("sh", "sz")):
        code = code[2:]
    for product in products:
        try:
            spec = market_data.SSE_ETF_OPTION_SPECS[product]
        except KeyError:
            continue
        if code == _position_code_key(spec.etf_symbol).lower():
            return product
    return None


def _contract_multiplier(product):
    try:
        return float(core.config.load_config(product).vol.contract_multiplier)
    except Exception:
        return 1.0


def _restore_template_history(frame, sheet_name, columns):
    if not TEMPLATE_REPORT_PATH.exists():
        return frame
    template = account_report._read_report_workbook(TEMPLATE_REPORT_PATH).get(sheet_name)
    if template is None or template.empty or "日期" not in template.columns:
        return frame
    template = template.reindex(columns=columns)
    template_dates = pd.to_datetime(
        template["日期"],
        errors="coerce",
    ).dt.strftime("%Y-%m-%d")
    existing_dates = pd.to_datetime(
        frame["日期"],
        errors="coerce",
    ).dt.strftime("%Y-%m-%d")
    remaining = frame.loc[~existing_dates.isin(set(template_dates.dropna()))]
    return _concat_exact([template, remaining], columns)


def _deduplicate_trade_frame(frame):
    if frame.empty:
        return frame
    trade_id_column = account_report.TRADE_COLUMNS[5]
    if trade_id_column not in frame.columns:
        return frame.drop_duplicates().reset_index(drop=True)
    has_trade_id = frame[trade_id_column].notna() & frame[trade_id_column].astype(str).ne("")
    identified = frame.loc[has_trade_id].drop_duplicates(
        subset=[trade_id_column],
        keep="last",
    )
    unidentified = frame.loc[~has_trade_id].drop_duplicates()
    return _concat_exact([identified, unidentified], frame.columns)


def _latest_account_report_path(out_dir, before_path=None):
    before_path = Path(before_path) if before_path is not None else None
    candidates = [
        path
        for path in Path(out_dir).glob("????????_??????_account_report.xlsx")
        if before_path is None or path != before_path
    ]
    return max(candidates, key=lambda path: path.name) if candidates else None


def _apply_template_layout(workbook):
    if not TEMPLATE_REPORT_PATH.exists():
        return
    from openpyxl import load_workbook

    template = load_workbook(TEMPLATE_REPORT_PATH)
    summary_source = (
        template[SUMMARY_TEMPLATE_SHEET]
        if SUMMARY_TEMPLATE_SHEET in template.sheetnames
        else None
    )
    for sheet_name in workbook.sheetnames:
        if sheet_name in DETAIL_SHEETS:
            if sheet_name not in template.sheetnames:
                continue
            source = template[sheet_name]
        else:
            if summary_source is None:
                continue
            source = summary_source
        if sheet_name not in workbook.sheetnames:
            continue
        target = workbook[sheet_name]
        target.freeze_panes = source.freeze_panes
        for column_index in range(1, source.max_column + 1):
            column_letter = source.cell(1, column_index).column_letter
            target.column_dimensions[column_letter].width = (
                source.column_dimensions[column_letter].width
            )


def _fmt(value):
    try:
        return f"{float(value):,.2f}"
    except (TypeError, ValueError):
        return str(value)
