from __future__ import annotations

import os
import math
import re
import unicodedata
from pathlib import Path

import numpy as np
import pandas as pd

import core
from . import account as account_store
from . import market_data
from . import signal_engine
from . import storage
from .runtime import PROJECT_ROOT, load_product_config


SUMMARY_COLUMNS = [
    "日期",
    "账户ID",
    "标的价格",
    "对冲持仓",
    "对冲最新价",
    "当日手续费",
    "期权单日盈亏",
    "ETF单日盈亏",
    "总单日盈亏",
    "净单日盈亏",
    "持仓盈亏",
    "交易盈亏",
    "当日盯市交易盈亏",
    "当日盈亏分解合计",
    "当日盈亏对账差额",
    "券商期权单日盈亏变化",
    "券商对冲单日盈亏变化",
    "券商总单日盈亏变化",
    "账户Delta",
    "期权Delta",
    "账户Gamma",
    "账户Vega",
    "账户Theta",
    "持仓IV",
    "Call IV",
    "Put IV",
    "Call Delta",
    "Put Delta",
    "Call Gamma",
    "Put Gamma",
    "Call Vega",
    "Put Vega",
    "Call Theta",
    "Put Theta",
    "期权单日DeltaPnL",
    "期权单日GammaPnL",
    "期权单日VegaPnL",
    "期权单日ThetaPnL",
    "期权单日GreeksPnL",
    "对冲单日DeltaPnL",
    "对冲单日GreeksPnL",
    "交易DeltaPnL",
    "交易GammaPnL",
    "交易VegaPnL",
    "交易ThetaPnL",
    "交易GreeksPnL",
    "昨仓GreeksPnL",
    "单日DeltaPnL",
    "单日GammaPnL",
    "单日VegaPnL",
    "单日ThetaPnL",
    "单日GreeksPnL",
    "GreeksPnL口径",
    "GreeksPnL说明",
    "GreeksPnL路径节点数",
]

REPORT_MODES = {"default", "diagnose"}

DEFAULT_SUMMARY_REPORT_COLUMNS = [
    "日期",
    "当日手续费",
    "期权单日盈亏",
    "ETF单日盈亏",
    "总单日盈亏(手续费前)",
    "净单日盈亏",
    "单日盈亏/AUM",
    "账户Delta",
    "账户Gamma",
    "账户Vega",
    "账户Theta",
    "单日DeltaPnL",
    "单日GammaPnL",
    "单日VegaPnL",
    "单日ThetaPnL",
    "单日GreeksPnL",
]

DIAGNOSE_SUMMARY_REPORT_COLUMNS = [
    "日期",
    "当日手续费",
    "期权单日盈亏",
    "ETF单日盈亏",
    "总单日盈亏(手续费前)",
    "净单日盈亏",
    "单日盈亏/AUM",
    "持仓盈亏",
    "交易盈亏",
    "当日盯市交易盈亏",
    "当日盈亏分解合计",
    "当日盈亏对账差额",
    "账户Delta",
    "账户Gamma",
    "账户Vega",
    "账户Theta",
    "单日DeltaPnL",
    "单日GammaPnL",
    "单日VegaPnL",
    "单日ThetaPnL",
    "单日GreeksPnL",
]

DEFAULT_POSITION_REPORT_COLUMNS = [
    "日期",
    "合约代码",
    "合约名称",
    "交易方向",
    "总持仓张数",
    "AUM",
    "今日变化",
    "最新价",
    "持仓均价",
    "持仓盈亏",
    "交易盈亏",
    "到期日",
    "IV",
]

DIAGNOSE_POSITION_REPORT_COLUMNS = [
    "日期",
    "合约代码",
    "合约名称",
    "交易方向",
    "总持仓张数",
    "AUM",
    "今日变化",
    "最新价",
    "持仓均价",
    "持仓盈亏",
    "交易盈亏",
    "当日盯市交易盈亏",
    "当日盈亏分解合计",
    "到期日",
    "IV",
    "单张Gamma",
    "单张Vega",
    "单张Theta",
]

INTERNAL_RECONCILIATION_COLUMNS = {
    "持仓盈亏",
    "当日盯市交易盈亏",
    "当日盈亏分解合计",
    "当日盈亏对账差额",
    "券商期权单日盈亏变化",
    "券商对冲单日盈亏变化",
    "券商总单日盈亏变化",
    "期权单日DeltaPnL",
    "期权单日GammaPnL",
    "期权单日VegaPnL",
    "期权单日ThetaPnL",
    "期权单日GreeksPnL",
    "对冲单日DeltaPnL",
    "对冲单日GreeksPnL",
    "交易DeltaPnL",
    "交易GammaPnL",
    "交易VegaPnL",
    "交易ThetaPnL",
    "交易GreeksPnL",
    "昨仓GreeksPnL",
    "GreeksPnL口径",
    "GreeksPnL说明",
    "GreeksPnL路径节点数",
}

DIAGNOSTIC_REPORT_COLUMNS = [
    "日期",
    "账户ID",
    "券商总单日盈亏变化",
    "单日GreeksPnL",
    "Greeks解释残差",
    "GreeksPnL口径",
    "GreeksPnL说明",
    "GreeksPnL路径节点数",
]

DAILY_GREEKS_PNL_COLUMNS = [
    "期权单日DeltaPnL",
    "期权单日GammaPnL",
    "期权单日VegaPnL",
    "期权单日ThetaPnL",
    "期权单日GreeksPnL",
    "对冲单日DeltaPnL",
    "对冲单日GreeksPnL",
    "交易DeltaPnL",
    "交易GammaPnL",
    "交易VegaPnL",
    "交易ThetaPnL",
    "交易GreeksPnL",
    "昨仓GreeksPnL",
    "单日DeltaPnL",
    "单日GammaPnL",
    "单日VegaPnL",
    "单日ThetaPnL",
    "单日GreeksPnL",
]

POSITION_COLUMNS = [
    "日期",
    "账户ID",
    "方向",
    "合约代码",
    "合约名称",
    "买卖",
    "持仓类型",
    "总持仓",
    "今持仓",
    "今开仓",
    "今平仓",
    "可平量",
    "最新价",
    "持仓均价",
    "开仓均价",
    "期权市值",
    "占用保证金",
    "持仓盈亏",
    "浮动盈亏",
    "AUM",
    "行权价",
    "到期日",
    "剩余天数",
    "IV",
    "单张Delta",
    "Delta",
    "Gamma",
    "Vega",
    "Theta",
]

TRADE_COLUMNS = [
    "序号",
    "投资者账号",
    "交易所",
    "合约代码",
    "合约名称",
    "成交编号",
    "报单编号",
    "开平",
    "买卖",
    "报单价格",
    "成交价格",
    "成交数量",
    "手续费",
    "平仓盈亏",
    "类型",
    "日期",
    "报单时间",
    "成交时间",
]

PRODUCT_CONTRACT_NAME_MARKERS = {
    "50etf": "50ETF",
    "300etf": "300ETF",
    "500etf": "500ETF",
    "kc50etf": "科创50",
}


def build_live_account_report(
    product,
    account_id="default",
    source="snapshot",
    date=None,
    all_trades=False,
    persist_history=True,
):
    market_data.require_live_product(product)
    _, snapshot, market, _ = prepare_account_report_market(
        product,
        source=source,
        date=date,
    )
    if persist_history:
        _ensure_account_report_history(
            product,
            account_id,
            through_date=str(market["date"].date()),
        )
    payload = calculate_live_account_report(
        product,
        account_id=account_id,
        source=source,
        snapshot=snapshot,
        market=market,
        all_trades=all_trades,
    )
    payload["position_rows"] = _repair_zero_iv_position_rows_with_intraday_minutes(
        payload.get("position_rows", []),
        product,
    )
    _refresh_current_summary_greeks_from_position_rows(payload)
    if persist_history:
        persist_account_report_history(product, account_id, payload)
    else:
        payload["position_history"] = _read_report_history_for_calculation(
            storage.account_report_position_history_path(product, account_id),
            payload.get("position_rows", []),
            POSITION_COLUMNS,
            key_columns=["日期", "账户ID"],
        )
        payload["position_history"] = _backfill_position_single_delta_columns(
            payload["position_history"],
            product=product,
        )
        payload["position_history"] = _revalue_stale_position_greeks(
            payload["position_history"],
            product,
        )
        payload["position_history"] = _repair_zero_iv_position_rows_with_intraday_minutes(
            payload["position_history"],
            product,
        )
        _apply_current_pnl_decomposition(payload)
        payload["summary_history"] = _read_report_history_for_calculation(
            storage.account_report_summary_history_path(product, account_id),
            [payload["summary"]],
            SUMMARY_COLUMNS,
            key_columns=["日期", "账户ID"],
        )
        payload["summary_history"] = _backfill_summary_financial_columns(
            product,
            account_id,
            payload["summary_history"],
        )
        payload["summary_history"] = _add_summary_greeks_pnl(
            payload["summary_history"],
            payload["position_history"],
            product=product,
            current_position_report=_position_report_frame(payload),
            trade_rows=payload.get("trade_rows"),
            account_id=account_id,
        )
        _apply_current_pnl_decomposition_to_history(payload)
        _refresh_current_summary_from_history(payload)
    return payload


def _read_report_history_for_calculation(path, new_rows, columns, key_columns):
    path = Path(path)
    if path.exists():
        history = pd.read_csv(path, encoding="utf-8-sig")
    else:
        history = pd.DataFrame(columns=columns)
    incoming = _frame(new_rows, columns)
    if incoming.empty:
        return history.reindex(columns=columns)
    for column in columns:
        if column not in history.columns:
            history[column] = None
    mask = pd.Series(False, index=history.index)
    for _, row in incoming.iterrows():
        row_mask = pd.Series(True, index=history.index)
        for key in key_columns:
            row_mask &= history[key].astype(str) == str(row[key])
        mask |= row_mask
    history = history.loc[~mask]
    if history.empty:
        return incoming.reindex(columns=columns)
    return _concat_rows([history, incoming], columns=columns)


def _ensure_account_report_history(product, account_id, through_date=None):
    summary_path = storage.account_report_summary_history_path(product, account_id)
    position_path = storage.account_report_position_history_path(product, account_id)
    summary_exists = summary_path.exists()
    position_exists = position_path.exists()
    if summary_exists and position_exists:
        return None
    if summary_exists != position_exists:
        raise RuntimeError(
            "Account report history is incomplete: summary and position history "
            "must either both exist or both be absent."
        )

    total_path = _latest_total_report_path(storage.output_dir(product), product)
    if total_path is None:
        return None
    reset_at = account_store.load_account(product, account_id=account_id).reset_at
    restore_account_report_history_from_total(
        product,
        account_id=account_id,
        total_path=total_path,
        from_date=_date_or_none(reset_at),
        through_date=through_date,
    )
    return total_path


def restore_account_report_history_from_total(
    product,
    account_id="default",
    total_path=None,
    from_date=None,
    through_date=None,
):
    total_path = (
        Path(total_path)
        if total_path is not None
        else _latest_total_report_path(storage.output_dir(product), product)
    )
    if total_path is None or not total_path.exists():
        raise FileNotFoundError("No cumulative account report is available for history restore.")

    frames = _read_report_workbook(total_path)
    summary_report = frames.get("账户总体情况")
    position_report = frames.get("持仓记录")
    if summary_report is None or position_report is None:
        raise ValueError("Cumulative account report is missing required sheets.")

    summary_report = _history_rows_between_dates(
        summary_report,
        from_date,
        through_date,
    )
    position_report = _history_rows_between_dates(
        position_report,
        from_date,
        through_date,
    )
    if summary_report.empty or position_report.empty:
        raise ValueError("Cumulative account report contains no restorable history rows.")

    position_history = _restore_position_history_from_total(
        product,
        account_id,
        position_report,
    )
    position_history = _revalue_stale_position_greeks(position_history, product)
    position_history = _repair_zero_iv_position_rows_with_intraday_minutes(
        position_history,
        product,
    )
    summary_history = _restore_summary_history_from_total(
        product,
        account_id,
        summary_report,
        position_history,
    )
    summary_history = _add_summary_greeks_pnl(
        summary_history,
        position_history,
        product=product,
    )

    summary_path = storage.account_report_summary_history_path(product, account_id)
    position_path = storage.account_report_position_history_path(product, account_id)
    position_history.to_csv(position_path, index=False, encoding="utf-8-sig")
    summary_history.to_csv(summary_path, index=False, encoding="utf-8-sig")
    return {
        "summary_path": summary_path,
        "position_path": position_path,
        "source_total": total_path,
    }


def _history_rows_between_dates(frame, from_date, through_date):
    result = frame.copy()
    if "日期" in result.columns:
        dates = pd.to_datetime(result["日期"], errors="coerce")
        mask = pd.Series(True, index=result.index)
        if from_date is not None:
            mask &= dates.ge(pd.Timestamp(from_date).normalize().tz_localize(None))
        if through_date is not None:
            mask &= dates.le(pd.Timestamp(through_date).normalize().tz_localize(None))
        result = result.loc[mask]
    return _sort_report_frame(result)


def _restore_position_history_from_total(product, account_id, position_report):
    config = load_product_config(product)
    multiplier = float(config.vol.contract_multiplier)
    spots = {}
    for _, report_row in position_report.iterrows():
        if pd.isna(report_row.get("到期日")):
            report_date = str(pd.Timestamp(report_row.get("日期")).date())
            spots[report_date] = _number(report_row.get("最新价"))
    restored = []
    for _, report_row in position_report.iterrows():
        report_row = report_row.copy()
        report_date = str(pd.Timestamp(report_row.get("日期")).date())
        report_row["_恢复标的价格"] = spots.get(report_date)
        is_option = pd.notna(report_row.get("到期日"))
        side = (
            "short"
            if str(report_row.get("交易方向")) == "空"
            else "long"
            if is_option
            else "hedge"
        )
        qty = float(_number(report_row.get("总持仓张数")) or 0.0)
        latest = _number(report_row.get("最新价"))
        cost = _number(report_row.get("持仓均价"))
        direction = -1.0 if side == "short" else 1.0
        row = {
            "日期": report_date,
            "账户ID": account_id,
            "方向": side,
            "合约代码": _security_code(report_row.get("合约代码")),
            "合约名称": report_row.get("合约名称"),
            "买卖": "卖" if direction < 0 else "买",
            "持仓类型": "ETF对冲" if side == "hedge" else "义务仓" if side == "short" else "权利仓",
            "总持仓": qty,
            "今持仓": None,
            "今开仓": None,
            "今平仓": None,
            "可平量": None,
            "最新价": latest,
            "持仓均价": cost,
            "开仓均价": cost,
            "期权市值": (
                None
                if latest is None
                else latest * qty * (multiplier if is_option else 1.0)
            ),
            "占用保证金": None,
            "持仓盈亏": report_row.get("持仓盈亏"),
            "浮动盈亏": (
                None
                if latest is None or cost is None
                else direction * qty * (latest - cost) * (multiplier if is_option else 1.0)
            ),
            "行权价": None,
            "到期日": (
                str(pd.Timestamp(report_row.get("到期日")).date())
                if is_option
                else None
            ),
            "剩余天数": None,
            "IV": report_row.get("IV") if is_option else None,
            "单张Delta": None,
            "Delta": qty if side == "hedge" else None,
            "Gamma": 0.0 if side == "hedge" else None,
            "Vega": 0.0 if side == "hedge" else None,
            "Theta": 0.0 if side == "hedge" else None,
        }
        if is_option:
            _restore_option_position_greeks(row, report_row, config)
        restored.append(row)
    return pd.DataFrame(restored, columns=POSITION_COLUMNS)


def _restore_option_position_greeks(row, report_row, config):
    iv = _number(report_row.get("IV"))
    spot = _restored_spot_for_date(report_row.get("日期"), report_row)
    strike = _strike_from_contract_name(report_row.get("合约名称"), spot)
    maturity = _date_or_none(report_row.get("到期日"))
    report_date = _date_or_none(report_row.get("日期"))
    option_type = _option_type_from_contract_name(report_row.get("合约名称"))
    qty = abs(_number(report_row.get("总持仓张数")) or 0.0)
    if None in {iv, spot, strike, maturity, report_date, option_type} or qty <= 0:
        raise ValueError(
            f"Cannot restore option history for contract {report_row.get('合约代码')}."
        )

    dte = core.vol_engine._count_trading_dte(report_date, maturity)
    chain = pd.DataFrame(
        [
            {
                "option_type": option_type,
                "pricing_spot": spot,
                "strike_price": strike,
                "ttm": dte / float(config.vol.annual_days),
                "iv": iv,
            }
        ]
    )
    greeks = core.vol_engine.add_greeks_for_day(chain, spot).iloc[0]
    direction = -1.0 if row["方向"] == "short" else 1.0
    scale = direction * qty * float(config.vol.contract_multiplier)
    row["行权价"] = strike
    row["剩余天数"] = dte
    row["单张Delta"] = direction * float(greeks["delta"])
    for metric in ["Delta", "Gamma", "Vega", "Theta"]:
        row[metric] = float(greeks[metric.lower()]) * scale


def _restored_spot_for_date(report_date, report_row):
    value = _number(report_row.get("_恢复标的价格"))
    if value is not None:
        return value
    raise ValueError(f"Missing ETF mark for restored report date {report_date}.")


def _strike_from_contract_name(name, spot=None):
    match = re.search(r"(\d{4,5})$", str(name or ""))
    if match is None:
        return None
    raw = float(match.group(1))
    if spot is None or spot <= 0:
        return raw
    candidates = [raw, raw / 10.0, raw / 100.0, raw / 1000.0]
    return min(candidates, key=lambda value: abs(np.log(value / spot)))


def _option_type_from_contract_name(name):
    text = str(name or "").upper()
    if "购" in text or "CALL" in text:
        return "c"
    if "沽" in text or "PUT" in text:
        return "p"
    return None


def _restore_summary_history_from_total(
    product,
    account_id,
    summary_report,
    position_history,
):
    config = load_product_config(product)
    initial_cash = float(config.backtest.initial_cash)
    positions = position_history.copy()
    spots = (
        positions.loc[positions["方向"].astype(str).eq("hedge"), ["日期", "最新价"]]
        .drop_duplicates("日期", keep="last")
        .set_index("日期")["最新价"]
        .to_dict()
    )
    option_realized = 0.0
    hedge_realized = 0.0
    cumulative_fee = 0.0
    previous_option_unrealized = 0.0
    previous_hedge_unrealized = 0.0
    rows = []
    for _, report_row in summary_report.iterrows():
        report_date = str(pd.Timestamp(report_row.get("日期")).date())
        date_positions = positions.loc[positions["日期"].astype(str).eq(report_date)]
        option_positions = date_positions.loc[
            ~date_positions["方向"].astype(str).eq("hedge")
        ]
        hedge_positions = date_positions.loc[
            date_positions["方向"].astype(str).eq("hedge")
        ]
        daily_fee = _number(report_row.get("当日手续费")) or 0.0
        cumulative_fee += daily_fee
        option_unrealized = _sum_numeric_column(option_positions, "浮动盈亏")
        hedge_unrealized = _sum_numeric_column(hedge_positions, "浮动盈亏")
        option_daily_pnl = _number(report_row.get("期权单日盈亏")) or 0.0
        hedge_daily_pnl = _number(report_row.get("ETF单日盈亏")) or 0.0
        option_realized += option_daily_pnl - (
            option_unrealized - previous_option_unrealized
        )
        hedge_realized += hedge_daily_pnl - (
            hedge_unrealized - previous_hedge_unrealized
        )
        previous_option_unrealized = option_unrealized
        previous_hedge_unrealized = hedge_unrealized
        hedge_qty = _sum_numeric_column(hedge_positions, "总持仓")
        hedge_cost = _weighted_position_value(hedge_positions, "持仓均价")
        spot = _number(spots.get(report_date))
        row = {column: None for column in SUMMARY_COLUMNS}
        row.update(
            {
                "日期": report_date,
                "账户ID": account_id,
                "初始资金": initial_cash,
                "标的价格": spot,
                "对冲持仓": hedge_qty,
                "对冲成本": hedge_cost,
                "对冲最新价": spot,
                "对冲估值价": spot,
                "对冲估值价类型": _hedge_mark_price_type(report_date),
                "对冲浮盈亏": hedge_unrealized,
                "对冲已实现盈亏": hedge_realized,
                "对冲总盈亏": hedge_unrealized + hedge_realized,
                "估算权益": _number(report_row.get("估算权益")),
                "期权浮盈亏": option_unrealized,
                "期权已实现盈亏": option_realized,
                "期权总盈亏": option_unrealized + option_realized,
                "手续费": cumulative_fee,
                "当日手续费": daily_fee,
                "期权单日盈亏": _number(report_row.get("期权单日盈亏")),
                "对冲单日盈亏": _number(report_row.get("ETF单日盈亏")),
                "ETF单日盈亏": _number(report_row.get("ETF单日盈亏")),
                "总单日盈亏": _number(report_row.get("总单日盈亏(手续费前)")),
                "净单日盈亏": _number(report_row.get("净单日盈亏")),
                "账户Delta": _number(report_row.get("账户Delta")),
                "期权Delta": (
                    None
                    if _number(report_row.get("账户Delta")) is None
                    else _number(report_row.get("账户Delta")) - hedge_qty
                ),
                "账户Gamma": _number(report_row.get("账户Gamma")),
                "账户Vega": _number(report_row.get("账户Vega")),
                "账户Theta": _number(report_row.get("账户Theta")),
            }
        )
        rows.append(row)
    return pd.DataFrame(rows, columns=SUMMARY_COLUMNS)


def prepare_account_report_market(product, source="snapshot", date=None):
    config = load_product_config(product)
    snapshot = None
    report_date = date
    if source in {"akshare", "local", "snapshot"}:
        snapshot = market_data.fetch_quote_snapshot(
            product,
            source=source,
            date=date or "latest",
        )
        report_date = snapshot["quote_date"]
    elif source != "none":
        raise ValueError("source must be one of: akshare, local, snapshot, none")

    market = signal_engine._load_market_context(
        config,
        report_date,
        quote_snapshot=snapshot,
    )
    return config, snapshot, market, report_date


def calculate_live_account_report(
    product,
    account_id="default",
    source="snapshot",
    snapshot=None,
    market=None,
    all_trades=False,
):
    if market is None:
        _, snapshot, market, _ = prepare_account_report_market(
            product,
            source=source,
            date=None,
        )
    config = load_product_config(product)
    live_account = account_store.load_account(product, account_id=account_id)
    reset_at = live_account.reset_at
    report_date_text = str(market["date"].date())
    spot = float(market["signal_row"]["close"])
    report_hedge = _hedge_for_report_date(
        live_account,
        product,
        account_id,
        report_date_text,
    )

    position_rows, account_greeks, option_value, option_margin, option_pnl = (
        _position_rows_from_account(
            live_account,
            market["chain_df"],
            report_date_text,
            account_id,
        )
    )
    hedge_rows = _hedge_rows_from_account(
        product,
        report_hedge,
        account_id,
        report_date_text,
        spot,
        prefer_spot_mark=source in {"akshare", "local", "snapshot"},
        not_before=reset_at,
    )
    position_rows.extend(hedge_rows)
    trade_rows = _trade_rows_from_export(product, report_date_text, not_before=reset_at)
    trade_rows.extend(
        _etf_trade_rows_from_export(product, report_date_text, not_before=reset_at)
    )
    daily_fee = _configured_daily_report_fee(
        product,
        account_id,
        report_date_text,
        trade_rows,
    )

    hedge_latest_price = (
        _number(hedge_rows[0].get("最新价")) if hedge_rows else None
    )
    summary_row = {
        "日期": report_date_text,
        "账户ID": account_id,
        "标的价格": spot,
        "对冲持仓": report_hedge.qty,
        "对冲最新价": hedge_latest_price,
        "当日手续费": daily_fee,
        "账户Delta": account_greeks["delta"] + report_hedge.qty,
        "期权Delta": account_greeks["delta"],
        "账户Gamma": account_greeks["gamma"],
        "账户Vega": account_greeks["vega"],
        "账户Theta": account_greeks["theta"],
        "持仓IV": account_greeks["position_iv"],
        "Call IV": account_greeks["call_iv"],
        "Put IV": account_greeks["put_iv"],
        "Call Delta": account_greeks["call_delta"],
        "Put Delta": account_greeks["put_delta"],
        "Call Gamma": account_greeks["call_gamma"],
        "Put Gamma": account_greeks["put_gamma"],
        "Call Vega": account_greeks["call_vega"],
        "Put Vega": account_greeks["put_vega"],
        "Call Theta": account_greeks["call_theta"],
        "Put Theta": account_greeks["put_theta"],
    }

    return {
        "product": product,
        "account_id": account_id,
        "date": report_date_text,
        "spot": spot,
        "source": source,
        "quote_snapshot": snapshot,
        "summary": summary_row,
        "summary_history": None,
        "position_history": None,
        "position_rows": position_rows,
        "trade_rows": trade_rows,
        "current_chain_metadata": _chain_metadata(market["chain_df"]),
        "strategy_state": live_account.strategy_state.to_dict(),
    }


def persist_account_report_history(product, account_id, payload):
    summary_path = storage.account_report_summary_history_path(product, account_id)
    position_path = storage.account_report_position_history_path(product, account_id)
    payload["position_history"] = _update_history_csv(
        position_path,
        payload.get("position_rows", []),
        POSITION_COLUMNS,
        key_columns=["日期", "账户ID"],
    )
    payload["position_history"] = _backfill_position_single_delta_columns(
        payload["position_history"],
        product=product,
    )
    payload["position_history"] = _revalue_stale_position_greeks(
        payload["position_history"],
        product,
    )
    payload["position_history"] = _repair_zero_iv_position_rows_with_intraday_minutes(
        payload["position_history"],
        product,
    )
    payload["position_history"].to_csv(position_path, index=False, encoding="utf-8-sig")
    _apply_current_pnl_decomposition(payload)
    payload["summary_history"] = _update_history_csv(
        summary_path,
        [payload["summary"]],
        SUMMARY_COLUMNS,
        key_columns=["日期", "账户ID"],
    )
    payload["summary_history"] = _backfill_summary_financial_columns(
        product,
        account_id,
        payload["summary_history"],
    )
    payload["summary_history"] = _add_summary_greeks_pnl(
        payload["summary_history"],
        payload["position_history"],
        product=product,
        current_position_report=_position_report_frame(payload),
        trade_rows=payload.get("trade_rows"),
        account_id=account_id,
    )
    _apply_current_pnl_decomposition_to_history(payload)
    payload["summary_history"].to_csv(summary_path, index=False, encoding="utf-8-sig")
    _refresh_current_summary_from_history(payload)
    return payload


def write_live_account_report(product, payload, mode="default"):
    _validate_report_mode(mode)
    stamp = storage.local_now_stamp()
    out_dir = storage.output_dir(product)
    frames = _daily_report_frames(payload, mode=mode)
    paths = {}
    name_suffix = "_diagnose" if mode == "diagnose" else ""

    total_path = out_dir / f"{stamp}_report{name_suffix}.xlsx"
    _append_daily_frames_to_total_report(
        total_path,
        frames,
        payload["date"],
        start_date=_report_history_start_date(payload),
        existing_path=_latest_total_report_path(
            out_dir,
            product,
            mode=mode,
            before_path=total_path,
        ),
    )
    paths["total_excel"] = total_path

    json_path = out_dir / f"{stamp}_daily{name_suffix}.json"
    storage.write_json(json_path, _json_payload(payload, mode=mode))
    paths["json"] = json_path
    return paths


def _daily_report_frames(payload, mode="default"):
    report_date = str(payload["date"])
    frames = _report_frames(payload, mode=mode)
    return {
        sheet_name: _rows_for_report_date(frame, report_date)
        for sheet_name, frame in frames.items()
    }


def _append_daily_frames_to_total_report(
    path,
    daily_frames,
    report_date,
    start_date=None,
    existing_path=None,
):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    source_path = Path(existing_path) if existing_path is not None else path
    existing = _read_report_workbook(source_path) if source_path.exists() else {}
    combined = {}
    for sheet_name, daily in daily_frames.items():
        daily = daily.copy()
        old = existing.get(sheet_name)
        if old is None and sheet_name == "交易记录":
            old = existing.get("当日交易记录")
        if old is None:
            old = pd.DataFrame(columns=daily.columns)
        old = old.copy()
        if sheet_name == "账户总体情况":
            if "备注" not in daily.columns:
                daily["备注"] = None
            if "备注" not in old.columns:
                old["备注"] = None
            existing_remarks = old.loc[
                _report_date_mask(old, report_date),
                ["日期", "备注"],
            ]
            if not existing_remarks.empty and daily["备注"].isna().all():
                daily.loc[:, "备注"] = existing_remarks.iloc[-1]["备注"]
        old = old.reindex(columns=daily.columns)
        if start_date is not None:
            old = old.loc[_report_date_on_or_after_mask(old, start_date)]
        old = old.loc[~_report_date_mask(old, report_date)]
        frame = _concat_rows([old, daily], columns=daily.columns)
        if start_date is not None:
            frame = frame.loc[_report_date_on_or_after_mask(frame, start_date)]
        combined[sheet_name] = _sort_report_frame(frame)
    temp_path = path.with_name(f"{path.stem}.{os.getpid()}.tmp{path.suffix}")
    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
        for sheet_name, frame in combined.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
        _format_account_report_workbook(writer.book)
    temp_path.replace(path)


def _format_account_report_workbook(workbook):
    from openpyxl.utils import get_column_letter

    money_headers = {
        "当日手续费",
        "期权单日盈亏",
        "ETF单日盈亏",
        "总单日盈亏",
        "总单日盈亏(手续费前)",
        "净单日盈亏",
        "持仓盈亏",
        "交易盈亏",
        "AUM",
        "当日盯市交易盈亏",
        "当日盈亏分解合计",
        "当日盈亏对账差额",
        "手续费",
        "平仓盈亏",
        "账户Delta",
        "账户Gamma",
        "账户Vega",
        "账户Theta",
        "单日DeltaPnL",
        "单日GammaPnL",
        "单日VegaPnL",
        "单日ThetaPnL",
        "单日GreeksPnL",
        "单张Delta",
        "单张Gamma",
        "单张Vega",
        "单张Theta",
    }
    price_headers = {"最新价", "持仓均价", "报单价格", "成交价格"}
    integer_headers = {"序号", "总持仓张数", "今日变化", "成交数量"}
    percent_headers = {"IV", "单日盈亏/AUM"}

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_column:
            last_column = get_column_letter(worksheet.max_column)
            worksheet.auto_filter.ref = f"A1:{last_column}{worksheet.max_row}"
        headers = {
            cell.column: str(cell.value or "")
            for cell in worksheet[1]
        }
        for column_index, header in headers.items():
            if header in money_headers:
                number_format = "#,##0.00;-#,##0.00;0.00"
            elif header in price_headers:
                number_format = "0.00000;-0.00000;0.00000"
            elif header in integer_headers:
                number_format = "#,##0;-#,##0;0"
            elif header in percent_headers:
                number_format = "0.00%;-0.00%;0.00%"
            else:
                number_format = None
            if number_format is not None:
                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.cell(row_index, column_index).number_format = number_format

            max_length = _excel_display_width(header)
            for row_index in range(2, worksheet.max_row + 1):
                value = worksheet.cell(row_index, column_index).value
                if value is not None:
                    max_length = max(max_length, _excel_display_width(value))
            if header == "备注":
                width = min(max(max_length + 2, 14), 40)
            elif header in {"成交编号", "投资者账号", "合约名称"}:
                width = min(max(max_length + 2, 14), 32)
            elif header == "总单日盈亏(手续费前)":
                width = 20
            else:
                width = min(max(max_length + 2, 11), 24)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        worksheet.row_dimensions[1].height = 22
        for row_index in range(2, worksheet.max_row + 1):
            line_count = max(
                str(worksheet.cell(row_index, column_index).value or "").count("\n") + 1
                for column_index in range(1, worksheet.max_column + 1)
            )
            worksheet.row_dimensions[row_index].height = max(18, line_count * 15)


def _excel_display_width(value):
    """Estimate the width Excel needs, accounting for full-width CJK text."""
    lines = str(value).splitlines() or [""]
    return max(
        sum(
            2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
            for character in line
        )
        for line in lines
    )


def _latest_total_report_path(out_dir, product, mode="default", before_path=None):
    out_dir = Path(out_dir)
    name_suffix = "_diagnose" if mode == "diagnose" else ""
    before_path = Path(before_path) if before_path is not None else None
    candidates = [
        path
        for path in out_dir.glob(f"????????_??????_report{name_suffix}.xlsx")
        if before_path is None or path != before_path
    ]
    if candidates:
        return max(candidates, key=lambda path: path.name)
    legacy_path = out_dir / f"{product}_account_report_total{name_suffix}.xlsx"
    return legacy_path if legacy_path.exists() else None


def _read_report_workbook(path):
    with pd.ExcelFile(path) as workbook:
        return {
            sheet_name: workbook.parse(sheet_name=sheet_name)
            for sheet_name in workbook.sheet_names
        }


def _rows_for_report_date(frame, report_date):
    if frame.empty or "日期" not in frame.columns:
        return frame.copy()
    return frame.loc[_report_date_mask(frame, report_date)].reset_index(drop=True)


def _report_date_mask(frame, report_date):
    if frame.empty or "日期" not in frame.columns:
        return pd.Series(False, index=frame.index)
    dates = pd.to_datetime(frame["日期"], errors="coerce").dt.strftime("%Y-%m-%d")
    return dates.eq(str(report_date))


def _report_date_on_or_after_mask(frame, start_date):
    if frame.empty or "日期" not in frame.columns:
        return pd.Series(True, index=frame.index)
    dates = pd.to_datetime(frame["日期"], errors="coerce")
    return dates.ge(pd.Timestamp(start_date).normalize())


def _report_history_start_date(payload):
    history = payload.get("summary_history")
    if not isinstance(history, pd.DataFrame) or history.empty or "日期" not in history:
        return payload.get("date")
    dates = pd.to_datetime(history["日期"], errors="coerce").dropna()
    return str(dates.min().date()) if not dates.empty else payload.get("date")


def _sort_report_frame(frame):
    if frame.empty or "日期" not in frame.columns:
        return frame.reset_index(drop=True)
    result = frame.copy()
    result["_report_sort_date"] = pd.to_datetime(result["日期"], errors="coerce")
    result = result.sort_values("_report_sort_date", kind="stable")
    return result.drop(columns=["_report_sort_date"]).reset_index(drop=True)


def _json_payload(payload, mode="default"):
    _validate_report_mode(mode)
    result = dict(payload)
    result.pop("current_chain_metadata", None)
    for key in ["summary_history", "position_history"]:
        value = result.get(key)
        if isinstance(value, pd.DataFrame):
            result[key] = value.to_dict("records")
    if mode == "default":
        result["summary"] = _without_internal_reconciliation_fields(
            result.get("summary")
        )
        result["summary_history"] = [
            _without_internal_reconciliation_fields(row)
            for row in result.get("summary_history", [])
        ]
    return result


def format_terminal_summary(payload, mode="default"):
    _validate_report_mode(mode)
    summary = payload["summary"]
    position_report = _position_report_frame(payload)
    pnl_decomposition = _position_pnl_totals(position_report)
    lines = [
        f"报告快照时间: {_snapshot_time_text(payload.get('quote_snapshot'))}",
        (
            f"账户报告={payload['product']}/{payload['account_id']} "
            f"模式={mode} 日期={payload['date']} 标的价格={_fmt(payload['spot'])}"
        ),
        (
            f"期权单日盈亏={_fmt(pnl_decomposition['option_daily_pnl'])} "
            f"ETF单日盈亏={_fmt(pnl_decomposition['etf_daily_pnl'])} "
            f"总单日盈亏(手续费前)={_fmt(pnl_decomposition['daily_pnl_decomposition'])} "
            f"净单日盈亏={_fmt(_net_daily_pnl(pnl_decomposition['daily_pnl_decomposition'], summary.get('当日手续费')))} "
            f"当日手续费={_fmt(summary.get('当日手续费'))}"
        ),
        (
            f"账户Delta={_fmt(summary['账户Delta'])} "
            f"Gamma={_fmt(summary['账户Gamma'])} "
            f"Vega={_fmt(summary['账户Vega'])} "
            f"Theta={_fmt(summary['账户Theta'])} "
            f"持仓IV={_fmt(summary['持仓IV'])}"
        ),
        (
            f"单日GreeksPnL={_fmt(summary.get('单日GreeksPnL'))} "
            f"Delta={_fmt(summary.get('单日DeltaPnL'))} "
            f"Gamma={_fmt(summary.get('单日GammaPnL'))} "
            f"Vega={_fmt(summary.get('单日VegaPnL'))} "
            f"Theta={_fmt(summary.get('单日ThetaPnL'))}"
        ),
    ]
    if mode == "diagnose":
        lines.extend(
            [
                (
                    f"持仓盈亏={_fmt(pnl_decomposition['holding_pnl'])} "
                    f"交易盈亏={_fmt(pnl_decomposition['realized_cost_pnl'])} "
                    f"当日盯市交易盈亏={_fmt(pnl_decomposition['mark_to_market_trade_pnl'])} "
                    f"当日盈亏分解合计={_fmt(pnl_decomposition['daily_pnl_decomposition'])}"
                ),
                (
                    f"券商差分总单日盈亏={_fmt(_broker_daily_pnl(summary))} "
                    f"对账差额={_fmt(_broker_reconciliation_difference(summary, pnl_decomposition))}"
                ),
            ]
        )
    lines.extend(["", "持仓记录"])
    lines.extend(
        _plain_table(
            position_report.to_dict("records"),
            ["交易方向", "合约代码", "合约名称", "总持仓张数", "今日变化", "最新价", "持仓均价", "IV"],
        )
    )
    lines.extend(["", "交易记录"])
    lines.extend(
        _plain_table(
            payload["trade_rows"],
            ["成交编号", "合约代码", "合约名称", "开平", "买卖", "成交价格", "成交数量", "成交时间"],
        )
    )
    return lines


def format_intraday_data_usage(payload, capture_result=None):
    product = payload.get("product")
    report_date = payload.get("date")
    lines = [
        (
            f"intraday_usage product={product} report_date={report_date} "
            f"valuation_source=quote_snapshot "
            f"snapshot_time={_snapshot_time_text(payload.get('quote_snapshot'))} "
            f"spot={_fmt(payload.get('spot'))}"
        )
    ]
    if capture_result is not None:
        errors = capture_result.get("errors") or []
        if errors:
            lines.append(
                f"WARNING intraday_capture product={product} errors={errors}"
            )
        else:
            lines.append(
                f"intraday_capture product={product} captured_at="
                f"{capture_result.get('captured_at')} etf_rows="
                f"{capture_result.get('etf_rows')} option_minute_rows="
                f"{capture_result.get('option_minute_rows')}"
            )

    lines.extend(_intraday_coverage_lines(product, report_date, payload))
    lines.extend(_transaction_price_source_lines(product, report_date, payload))
    return lines


def _intraday_coverage_lines(product, report_date, payload):
    if not product or not report_date:
        return []
    spec = market_data.SSE_ETF_OPTION_SPECS.get(product)
    if spec is None:
        return [f"intraday_coverage product={product} unsupported_product=true"]
    day = pd.Timestamp(report_date).strftime("%Y%m%d")
    root = storage.PROJECT_ROOT / "data" / "live" / product / "intraday" / day
    lines = []
    etf_path = root / f"etf_{spec.etf_symbol}_1m.csv"
    lines.append(
        _format_intraday_file_coverage(
            product,
            f"etf={spec.etf_symbol}",
            etf_path,
        )
    )
    for code in _intraday_option_codes_from_payload(payload):
        lines.append(
            _format_intraday_file_coverage(
                product,
                f"option={code}",
                root / f"option_{code}_1m.csv",
            )
        )
    return lines


def _intraday_option_codes_from_payload(payload):
    codes = []
    for row in payload.get("position_rows") or []:
        code = _security_code(row.get("合约代码"))
        if code and code.startswith("100"):
            codes.append(code)
    for row in payload.get("trade_rows") or []:
        code = _security_code(row.get("合约代码"))
        if code and code.startswith("100"):
            codes.append(code)
    return sorted(dict.fromkeys(codes))


def _format_intraday_file_coverage(product, label, path):
    summary = _intraday_file_coverage(path)
    if summary is None:
        return f"intraday_coverage product={product} {label} status=missing path={path}"
    return (
        f"intraday_coverage product={product} {label} rows={summary['rows']} "
        f"first={_fmt_timestamp(summary['first'])} "
        f"last={_fmt_timestamp(summary['last'])} path={path}"
    )


def _intraday_file_coverage(path):
    path = Path(path)
    if not path.exists():
        return None
    try:
        frame = pd.read_csv(path, encoding="utf-8-sig")
    except Exception:
        return None
    if frame.empty or "timestamp" not in frame.columns:
        return None
    timestamp = pd.to_datetime(frame["timestamp"], errors="coerce").dropna()
    if timestamp.empty:
        return None
    return {
        "rows": int(len(frame)),
        "first": timestamp.min(),
        "last": timestamp.max(),
    }


def _transaction_price_source_lines(product, report_date, payload):
    lines = []
    if not product or not report_date:
        return lines
    for row in payload.get("trade_rows") or []:
        if str(row.get("类型") or "") == "ETF对冲":
            continue
        code = _security_code(row.get("合约代码"))
        if not code or not code.startswith("100"):
            continue
        timestamp = _trade_row_timestamp(row, report_date)
        if timestamp is None:
            lines.append(
                f"transaction_price_source product={product} code={code} "
                "trade_time=missing spot_source=report_close "
                f"spot={_fmt(payload.get('spot'))}"
            )
            continue
        detail = _spot_from_intraday_minute_detail(product, report_date, timestamp)
        if detail is None:
            detail = _spot_from_quote_snapshot_detail(product, report_date, timestamp)
        if detail is None:
            lines.append(
                f"transaction_price_source product={product} code={code} "
                f"trade_time={_fmt_timestamp(timestamp)} "
                f"spot_source=report_close spot_time={report_date} "
                f"spot={_fmt(payload.get('spot'))}"
            )
            continue
        lines.append(
            f"transaction_price_source product={product} code={code} "
            f"trade_time={_fmt_timestamp(timestamp)} "
            f"spot_source={detail['source']} "
            f"spot_time={_fmt_timestamp(detail['timestamp'])} "
            f"spot={_fmt(detail['price'])}"
        )
    return lines


def _fmt_timestamp(value):
    if value is None:
        return "None"
    timestamp = pd.Timestamp(value)
    if pd.isna(timestamp):
        return "None"
    return timestamp.strftime("%Y-%m-%d %H:%M:%S")


def _snapshot_time_text(snapshot):
    if not snapshot:
        return "不可用"
    stamp = str(snapshot.get("snapshot_stamp") or "")
    timestamp = pd.to_datetime(stamp, format="%Y%m%d_%H%M%S", errors="coerce")
    if pd.isna(timestamp):
        return stamp or "未知"
    return timestamp.strftime("%Y-%m-%d %H:%M:%S")


def _report_frames(payload, mode="default"):
    _validate_report_mode(mode)
    position_report = _position_report_frame(payload)
    return {
        "账户总体情况": _summary_report_frame(
            payload["summary_history"],
            position_report=position_report,
            aum_by_date=_summary_aum_by_date(payload),
            report_date=payload.get("date"),
            mode=mode,
        ),
        "持仓记录": position_report.reindex(
            columns=(
                DIAGNOSE_POSITION_REPORT_COLUMNS
                if mode == "diagnose"
                else DEFAULT_POSITION_REPORT_COLUMNS
            )
        ),
        "交易记录": _frame(payload["trade_rows"], TRADE_COLUMNS),
    }


def _summary_report_frame(
    summary_history,
    position_report=None,
    aum_by_date=None,
    report_date=None,
    mode="default",
):
    _validate_report_mode(mode)
    report_columns = (
        DIAGNOSE_SUMMARY_REPORT_COLUMNS
        if mode == "diagnose"
        else DEFAULT_SUMMARY_REPORT_COLUMNS
    )
    if summary_history is None:
        return pd.DataFrame(columns=report_columns)
    frame = summary_history.copy()
    frame["期权单日盈亏"] = _prefer_numeric_column(
        frame,
        "期权单日盈亏",
        "券商期权单日盈亏变化",
    )
    frame["ETF单日盈亏"] = _prefer_numeric_column(
        frame,
        "ETF单日盈亏",
        "券商对冲单日盈亏变化",
    )
    frame["总单日盈亏"] = _prefer_numeric_column(
        frame,
        "总单日盈亏",
        "券商总单日盈亏变化",
    )
    frame["总单日盈亏(手续费前)"] = frame["总单日盈亏"]
    frame["净单日盈亏"] = pd.to_numeric(
        frame["总单日盈亏"],
        errors="coerce",
    ) - pd.to_numeric(frame.get("当日手续费"), errors="coerce").fillna(0.0)
    frame["单日盈亏/AUM"] = _daily_pnl_aum_ratio(
        frame,
        aum_by_date=aum_by_date,
        position_report=position_report,
    )
    if isinstance(position_report, pd.DataFrame) and not position_report.empty:
        report_date = str(report_date)
        current_mask = frame["日期"].astype(str).eq(report_date)
        totals = _position_pnl_totals(position_report)
        actual_total = pd.to_numeric(
            frame.loc[current_mask, "总单日盈亏"],
            errors="coerce",
        )
        frame.loc[current_mask, "持仓盈亏"] = totals["holding_pnl"]
        frame.loc[current_mask, "交易盈亏"] = totals["realized_cost_pnl"]
        frame.loc[current_mask, "当日盯市交易盈亏"] = totals[
            "mark_to_market_trade_pnl"
        ]
        frame.loc[current_mask, "当日盈亏分解合计"] = totals[
            "daily_pnl_decomposition"
        ]
        frame.loc[current_mask, "当日盈亏对账差额"] = (
            actual_total - totals["daily_pnl_decomposition"]
        )
    return frame.reindex(columns=report_columns)


def _summary_aum_by_date(payload):
    positions = payload.get("position_history")
    if not isinstance(positions, pd.DataFrame):
        positions = _frame(payload.get("position_rows", []), POSITION_COLUMNS)
    required = {"日期", "总持仓", "到期日"}
    if positions.empty or not required.issubset(positions.columns):
        return {}

    option_rows = positions.loc[positions["到期日"].notna()].copy()
    if option_rows.empty:
        return {}
    option_rows["_aum_date"] = pd.to_datetime(
        option_rows["日期"], errors="coerce"
    ).dt.strftime("%Y-%m-%d")
    option_rows["_aum_qty"] = pd.to_numeric(
        option_rows["总持仓"], errors="coerce"
    ).fillna(0.0).abs()
    group_columns = ["_aum_date"] + [
        column
        for column in ("账户ID", "方向", "行权价", "到期日")
        if column in option_rows.columns
    ]
    capacity = (
        option_rows.groupby(group_columns, dropna=False)["_aum_qty"]
        .max()
        .groupby(level=0)
        .sum()
    )

    summary = payload.get("summary_history")
    spots = {}
    if isinstance(summary, pd.DataFrame) and {"日期", "标的价格"}.issubset(
        summary.columns
    ):
        spot_frame = summary[["日期", "标的价格"]].copy()
        spot_frame["_aum_date"] = pd.to_datetime(
            spot_frame["日期"], errors="coerce"
        ).dt.strftime("%Y-%m-%d")
        spot_frame["_aum_spot"] = pd.to_numeric(
            spot_frame["标的价格"], errors="coerce"
        )
        spots = (
            spot_frame.dropna(subset=["_aum_date", "_aum_spot"])
            .drop_duplicates("_aum_date", keep="last")
            .set_index("_aum_date")["_aum_spot"]
            .to_dict()
        )
    report_date = str(payload.get("date"))
    current_spot = _number(payload.get("spot"))
    if current_spot is not None:
        spots.setdefault(report_date, current_spot)

    multiplier = _contract_multiplier(payload.get("product"))
    return {
        date: float(qty) * float(multiplier) * float(spots[date])
        for date, qty in capacity.items()
        if date in spots and float(qty) > 0
    }


def _daily_pnl_aum_ratio(frame, aum_by_date=None, position_report=None):
    dates = pd.to_datetime(frame["日期"], errors="coerce").dt.strftime("%Y-%m-%d")
    aum_lookup = dict(aum_by_date or {})
    if isinstance(position_report, pd.DataFrame) and not position_report.empty:
        for date, aum in _position_report_aum_by_date(position_report).items():
            aum_lookup.setdefault(date, aum)
    aum = pd.to_numeric(dates.map(aum_lookup), errors="coerce")
    gross_pnl = pd.to_numeric(
        frame["总单日盈亏(手续费前)"],
        errors="coerce",
    )
    return (gross_pnl / aum).where(aum.gt(0))


def _position_report_aum_by_date(position_report):
    required = {"日期", "AUM", "到期日"}
    if not required.issubset(position_report.columns):
        return {}
    option_rows = position_report.loc[position_report["到期日"].notna()].copy()
    if option_rows.empty:
        return {}
    option_rows["_aum_date"] = pd.to_datetime(
        option_rows["日期"], errors="coerce"
    ).dt.strftime("%Y-%m-%d")
    option_rows["_aum_value"] = pd.to_numeric(
        option_rows["AUM"], errors="coerce"
    )
    option_rows = option_rows.dropna(subset=["_aum_date", "_aum_value"])
    group_columns = ["_aum_date"] + [
        column for column in ("交易方向", "到期日") if column in option_rows.columns
    ]
    return (
        option_rows.groupby(group_columns, dropna=False)["_aum_value"]
        .max()
        .groupby(level=0)
        .sum()
        .to_dict()
    )


def _apply_current_pnl_decomposition(payload):
    totals = _position_pnl_totals(_position_report_frame(payload))
    option_daily_pnl = totals["option_daily_pnl"]
    etf_daily_pnl = totals["etf_daily_pnl"]
    total_daily_pnl = option_daily_pnl + etf_daily_pnl
    daily_fee = _number(payload["summary"].get("当日手续费")) or 0.0
    payload["summary"].update(
        {
            "期权单日盈亏": option_daily_pnl,
            "ETF单日盈亏": etf_daily_pnl,
            "总单日盈亏": total_daily_pnl,
            "净单日盈亏": total_daily_pnl - daily_fee,
            "持仓盈亏": totals["holding_pnl"],
            "交易盈亏": totals["realized_cost_pnl"],
            "当日盯市交易盈亏": totals["mark_to_market_trade_pnl"],
            "当日盈亏分解合计": totals["daily_pnl_decomposition"],
            "当日盈亏对账差额": total_daily_pnl - totals["daily_pnl_decomposition"],
        }
    )
    return totals


def _prefer_numeric_column(frame, preferred, fallback):
    preferred_values = (
        pd.to_numeric(frame[preferred], errors="coerce")
        if preferred in frame.columns
        else pd.Series(np.nan, index=frame.index)
    )
    fallback_values = (
        pd.to_numeric(frame[fallback], errors="coerce")
        if fallback in frame.columns
        else pd.Series(np.nan, index=frame.index)
    )
    return preferred_values.combine_first(fallback_values)


def _position_pnl_totals(position_report):
    holding_pnl = _sum_numeric_column(position_report, "持仓盈亏")
    transaction_pnl = _sum_numeric_column(position_report, "交易盈亏")
    mark_to_market_trade_pnl = _sum_numeric_column(
        position_report,
        "当日盯市交易盈亏",
    )
    daily_by_row = (
        pd.to_numeric(position_report.get("当日盈亏分解合计"), errors="coerce")
        .fillna(0.0)
    )
    option_mask = position_report.get("到期日").notna()
    option_daily_pnl = float(daily_by_row.loc[option_mask].sum())
    etf_daily_pnl = float(daily_by_row.loc[~option_mask].sum())
    return {
        "holding_pnl": holding_pnl,
        "realized_cost_pnl": transaction_pnl,
        "mark_to_market_trade_pnl": mark_to_market_trade_pnl,
        "daily_pnl_decomposition": holding_pnl + transaction_pnl,
        "option_daily_pnl": option_daily_pnl,
        "etf_daily_pnl": etf_daily_pnl,
    }


def _sum_numeric_column(frame, column):
    if column not in frame.columns:
        return 0.0
    return float(pd.to_numeric(frame[column], errors="coerce").fillna(0.0).sum())


def _net_daily_pnl(gross_pnl, daily_fee):
    if isinstance(daily_fee, pd.Series):
        return float(gross_pnl) - pd.to_numeric(
            daily_fee,
            errors="coerce",
        ).fillna(0.0)
    return float(gross_pnl) - float(_number(daily_fee) or 0.0)


def _apply_current_pnl_decomposition_to_history(payload):
    history = payload.get("summary_history")
    if history is None or history.empty:
        return
    mask = (
        history["日期"].astype(str).eq(str(payload["date"]))
        & history["账户ID"].astype(str).eq(str(payload["account_id"]))
    )
    if not mask.any():
        return

    totals = _position_pnl_totals(_position_report_frame(payload))
    row = history.loc[mask].iloc[-1]
    option_daily_pnl = totals["option_daily_pnl"]
    etf_daily_pnl = totals["etf_daily_pnl"]
    total_daily_pnl = option_daily_pnl + etf_daily_pnl
    daily_fee = _number(row.get("当日手续费")) or 0.0
    values = {
        "期权单日盈亏": option_daily_pnl,
        "ETF单日盈亏": etf_daily_pnl,
        "总单日盈亏": total_daily_pnl,
        "净单日盈亏": total_daily_pnl - daily_fee,
        "持仓盈亏": totals["holding_pnl"],
        "交易盈亏": totals["realized_cost_pnl"],
        "当日盯市交易盈亏": totals["mark_to_market_trade_pnl"],
        "当日盈亏分解合计": totals["daily_pnl_decomposition"],
        "当日盈亏对账差额": total_daily_pnl - totals["daily_pnl_decomposition"],
    }
    for column, value in values.items():
        history.loc[mask, column] = value


def _validate_report_mode(mode):
    if mode not in REPORT_MODES:
        raise ValueError("mode must be one of: default, diagnose")


def _without_internal_reconciliation_fields(row):
    if not isinstance(row, dict):
        return row
    return {
        key: value
        for key, value in row.items()
        if key not in INTERNAL_RECONCILIATION_COLUMNS
    }


def _broker_reconciliation_difference(summary, pnl_decomposition):
    broker_total = _broker_daily_pnl(summary)
    if broker_total is None:
        return None
    return broker_total - pnl_decomposition["daily_pnl_decomposition"]


def _broker_daily_pnl(summary):
    if str(summary.get("GreeksPnL说明") or "") == "first_history_row":
        return None
    return _number(summary.get("券商总单日盈亏变化"))


def _position_report_frame(payload):
    history = payload.get("position_history")
    if not isinstance(history, pd.DataFrame):
        history = _frame(payload.get("position_rows", []), POSITION_COLUMNS)
    if history.empty:
        return pd.DataFrame(columns=DIAGNOSE_POSITION_REPORT_COLUMNS)

    report_date = str(payload["date"])
    dates = history["日期"].astype(str)
    current = history.loc[dates.eq(report_date)].copy()
    prior_dates = sorted(date for date in dates.unique() if date < report_date)
    previous = (
        history.loc[dates.eq(prior_dates[-1])].copy()
        if prior_dates
        else pd.DataFrame(columns=history.columns)
    )
    today_trades = [
        row
        for row in payload.get("trade_rows", [])
        if str(row.get("日期")) == report_date
    ]
    trade_codes = {
        _security_code(row.get("合约代码"))
        for row in today_trades
        if row.get("合约代码") is not None and not pd.isna(row.get("合约代码"))
    }
    current_code_series = current["合约代码"].apply(_security_code)
    previous_code_series = previous["合约代码"].apply(_security_code)
    current_codes = set(current_code_series.dropna())
    previous_codes = set(previous_code_series.dropna())
    codes = sorted(current_codes | (previous_codes & trade_codes) | trade_codes)
    current_aum_by_code = _position_aum_by_code(payload, current)

    rows = []
    for code in codes:
        current_rows = current.loc[current_code_series.eq(code)]
        previous_rows = previous.loc[previous_code_series.eq(code)]
        trade_rows = [
            row for row in today_trades if _security_code(row.get("合约代码")) == code
        ]
        if current_rows.empty and previous_rows.empty and not trade_rows:
            continue
        rows.append(
            _position_report_row(
                payload,
                code,
                current_rows,
                previous_rows,
                trade_rows,
                current_aum_by_code,
            )
        )
    return pd.DataFrame(rows, columns=DIAGNOSE_POSITION_REPORT_COLUMNS)


def _position_report_row(
    payload,
    code,
    current_rows,
    previous_rows,
    trade_rows,
    current_aum_by_code=None,
):
    reference = (
        current_rows.iloc[0]
        if not current_rows.empty
        else previous_rows.iloc[0]
        if not previous_rows.empty
        else {}
    )
    side = reference.get("方向") if hasattr(reference, "get") else None
    if side is None:
        side = _side_from_trade_rows(trade_rows)
    current_qty = _rows_total_qty(current_rows)
    previous_qty = _rows_total_qty(previous_rows)
    direction_sign = _position_direction_sign(side, current_qty, previous_qty)
    metadata = payload.get("current_chain_metadata", {}).get(code, {})
    is_option = bool(metadata)
    latest_price = _option_mark_from_metadata(metadata) if is_option else None
    if latest_price is None and str(side).lower() == "hedge":
        latest_price = payload.get("spot")
    if latest_price is None:
        latest_price = _first_value(current_rows, "最新价")
    if latest_price is None:
        latest_price = _first_value(previous_rows, "最新价")

    cost_rows = current_rows if current_qty != 0 else previous_rows
    holding_cost = _weighted_position_value(cost_rows, "持仓均价")
    if holding_cost is None:
        holding_cost = _weighted_trade_open_price(trade_rows)
    multiplier = _number(metadata.get("contract_multiplier")) if is_option else 1.0
    pnl_breakdown = _daily_position_pnl_breakdown(
        current_qty=current_qty,
        current_side=side,
        current_price=latest_price,
        previous_qty=previous_qty,
        previous_side=_first_value(previous_rows, "方向"),
        previous_price=_first_value(previous_rows, "最新价"),
        previous_cost=_weighted_position_value(previous_rows, "持仓均价"),
        trade_rows=trade_rows,
        multiplier=multiplier or 1.0,
    )
    if pnl_breakdown["ending_cost"] is not None and current_qty != 0:
        holding_cost = pnl_breakdown["ending_cost"]
    contract_name = metadata.get("contract_symbol")
    if contract_name is None:
        contract_name = _first_value(current_rows, "合约名称")
    if contract_name is None:
        contract_name = _first_value(previous_rows, "合约名称")
    if contract_name is None and trade_rows:
        contract_name = trade_rows[0].get("合约名称")

    metadata_iv = _number(metadata.get("iv")) if is_option else None
    current_iv = _number(_first_value(current_rows, "IV"))
    use_current_greeks = (
        is_option
        and current_iv is not None
        and current_iv > 0
        and (metadata_iv is None or metadata_iv <= 0)
    )
    if is_option and use_current_greeks:
        single_delta = _number(_first_value(current_rows, "单张Delta"))
        single_gamma = _single_greek_from_position_rows(
            current_rows,
            "Gamma",
            current_qty,
            multiplier or 1.0,
        )
        single_vega = _single_greek_from_position_rows(
            current_rows,
            "Vega",
            current_qty,
            multiplier or 1.0,
        )
        single_theta = _single_greek_from_position_rows(
            current_rows,
            "Theta",
            current_qty,
            multiplier or 1.0,
        )
    elif is_option:
        single_delta = _signed_number(metadata.get("delta"), direction_sign)
        single_gamma = _signed_number(metadata.get("gamma"), direction_sign)
        single_vega = _signed_number(metadata.get("vega"), direction_sign)
        single_theta = _signed_number(metadata.get("theta"), direction_sign)
    else:
        single_delta = direction_sign
        single_gamma = 0.0
        single_vega = 0.0
        single_theta = 0.0

    return {
        "日期": payload["date"],
        "合约代码": code,
        "合约名称": contract_name,
        "交易方向": "空" if direction_sign < 0 else "多",
        "总持仓张数": current_qty,
        "今日变化": current_qty - previous_qty,
        "最新价": latest_price,
        "持仓均价": holding_cost,
        "持仓盈亏": pnl_breakdown["holding_pnl"],
        "交易盈亏": pnl_breakdown["realized_cost_pnl"],
        "当日盯市交易盈亏": pnl_breakdown["mark_to_market_trade_pnl"],
        "当日盈亏分解合计": pnl_breakdown["daily_pnl_decomposition"],
        "AUM": (
            (current_aum_by_code or {}).get(code)
            if (current_aum_by_code or {}).get(code) is not None
            else _first_value(current_rows, "AUM")
        ),
        "到期日": metadata.get("maturity_date") if is_option else None,
        "IV": current_iv if use_current_greeks else metadata.get("iv") if is_option else None,
        "单张Delta": single_delta,
        "单张Gamma": single_gamma,
        "单张Vega": single_vega,
        "单张Theta": single_theta,
    }


def _single_greek_from_position_rows(rows, column, current_qty, multiplier):
    value = _number(_first_value(rows, column))
    scale = abs(float(current_qty or 0.0)) * float(multiplier or 0.0)
    if value is None or scale <= 0:
        return None
    return value / scale


def _position_aum_by_code(payload, rows):
    if rows.empty or "合约代码" not in rows.columns:
        return {}
    spot = _number(payload.get("spot"))
    if spot is None:
        return {}
    multiplier = _contract_multiplier(payload.get("product"))
    option_rows = rows.copy()
    if "到期日" not in option_rows.columns or "行权价" not in option_rows.columns:
        return {}
    option_rows = option_rows.loc[
        option_rows["到期日"].notna() & option_rows["行权价"].notna()
    ].copy()
    if option_rows.empty:
        return {}
    option_rows["_qty_abs"] = (
        pd.to_numeric(option_rows["总持仓"], errors="coerce").fillna(0.0).abs()
    )
    result = {}
    for _, group in option_rows.groupby(
        ["账户ID", "方向", "行权价", "到期日"],
        dropna=False,
        sort=False,
    ):
        aum = float(group["_qty_abs"].max()) * float(multiplier) * float(spot)
        for code in group["合约代码"]:
            result[_security_code(code)] = aum
    return result


def _daily_position_pnl_breakdown(
    current_qty,
    current_side,
    current_price,
    previous_qty,
    previous_side,
    previous_price,
    previous_cost,
    trade_rows,
    multiplier,
):
    current_price = _number(current_price)
    previous_price = _number(previous_price)
    previous_cost = _number(previous_cost)
    multiplier = float(_number(multiplier) or 1.0)
    previous_signed_qty = _signed_position_qty(previous_qty, previous_side)
    total_signed_qty = previous_signed_qty
    total_cost = previous_cost
    transaction_pnl = 0.0
    holding_pnl = 0.0

    if current_price is not None and previous_price is not None:
        holding_pnl = (
            previous_signed_qty
            * (float(current_price) - float(previous_price))
            * multiplier
        )

    for trade in sorted(trade_rows, key=_trade_sort_key):
        trade_price = _number(trade.get("成交价格"))
        trade_signed_qty = _trade_signed_position_qty(trade)
        if trade_price is None or trade_signed_qty == 0:
            continue

        if current_price is not None:
            transaction_pnl += (
                trade_signed_qty
                * (float(current_price) - float(trade_price))
                * multiplier
            )

        remaining_trade_qty = trade_signed_qty
        if total_signed_qty != 0 and total_signed_qty * trade_signed_qty < 0:
            close_qty = min(abs(total_signed_qty), abs(trade_signed_qty))
            total_signed_qty -= np.sign(total_signed_qty) * close_qty
            remaining_trade_qty += np.sign(trade_signed_qty) * -close_qty
            if abs(total_signed_qty) <= 1e-9:
                total_signed_qty = 0.0
                total_cost = None

        if abs(remaining_trade_qty) > 1e-9:
            total_cost = _weighted_signed_cost(
                total_signed_qty,
                total_cost,
                remaining_trade_qty,
                trade_price,
            )
            total_signed_qty += remaining_trade_qty

    expected_signed_qty = _signed_position_qty(current_qty, current_side)
    ending_cost = total_cost if abs(total_signed_qty - expected_signed_qty) <= 1e-6 else None
    return {
        "holding_pnl": holding_pnl,
        "realized_cost_pnl": transaction_pnl,
        "mark_to_market_trade_pnl": transaction_pnl,
        "daily_pnl_decomposition": holding_pnl + transaction_pnl,
        "ending_cost": ending_cost,
    }


def _signed_position_qty(qty, side):
    qty = float(_number(qty) or 0.0)
    return -abs(qty) if str(side or "").lower() == "short" else qty


def _trade_signed_position_qty(trade):
    qty = float(_number(trade.get("成交数量")) or 0.0)
    return -qty if "卖" in str(trade.get("买卖") or "") else qty


def _weighted_signed_cost(current_qty, current_cost, added_qty, added_price):
    if abs(added_qty) <= 1e-9:
        return current_cost
    if abs(current_qty) <= 1e-9 or current_cost is None:
        return float(added_price)
    if current_qty * added_qty <= 0:
        return current_cost
    return (
        abs(current_qty) * float(current_cost) + abs(added_qty) * float(added_price)
    ) / (abs(current_qty) + abs(added_qty))


def _trade_sort_key(row):
    return str(
        row.get("成交时间(日)")
        or row.get("成交时间")
        or row.get("报单时间")
        or row.get("成交编号")
        or ""
    )


def _rows_total_qty(rows):
    if rows.empty or "总持仓" not in rows.columns:
        return 0.0
    return float(pd.to_numeric(rows["总持仓"], errors="coerce").fillna(0.0).sum())


def _first_value(rows, column):
    if rows.empty or column not in rows.columns:
        return None
    values = rows[column].dropna()
    return values.iloc[0] if not values.empty else None


def _weighted_position_value(rows, column):
    if rows.empty or column not in rows.columns:
        return None
    values = pd.to_numeric(rows[column], errors="coerce")
    qty = pd.to_numeric(rows["总持仓"], errors="coerce").abs()
    valid = values.notna() & qty.gt(0)
    if not valid.any():
        return _number(_first_value(rows, column))
    return float((values[valid] * qty[valid]).sum() / qty[valid].sum())


def _weighted_trade_open_price(rows):
    open_rows = [
        row
        for row in rows
        if "开" in str(row.get("开平") or "")
        and (_number(row.get("成交数量")) or 0.0) > 0
    ]
    total_qty = sum(_number(row.get("成交数量")) or 0.0 for row in open_rows)
    if total_qty <= 0:
        return None
    return sum(
        (_number(row.get("成交价格")) or 0.0) * (_number(row.get("成交数量")) or 0.0)
        for row in open_rows
    ) / total_qty


def _side_from_trade_rows(rows):
    for row in rows:
        if "开" not in str(row.get("开平") or ""):
            continue
        return "short" if "卖" in str(row.get("买卖") or "") else "long"
    return "hedge"


def _position_direction_sign(side, current_qty, previous_qty):
    if str(side or "").lower() == "short":
        return -1.0
    if str(side or "").lower() == "hedge" and (current_qty < 0 or previous_qty < 0):
        return -1.0
    return 1.0


def _signed_number(value, direction_sign):
    value = _number(value)
    return None if value is None else direction_sign * value


def _position_rows_from_account(live_account, chain_df, report_date, account_id):
    rows = []
    greeks_list = []
    signed_values = []
    margins = []
    option_pnl = 0.0
    for side, position in live_account.positions.items():
        if position is None:
            continue
        try:
            call_row, put_row = core.vol_engine.resolve_position_pair(position, chain_df)
        except IndexError:
            continue

        current_value = core.position.value(position, call_row, put_row)
        signed_value = core.position.signed_value(position, call_row, put_row)
        entry_value = float(position.get("entry_option_value", 0.0) or 0.0)
        pnl = entry_value - current_value if side == "short" else current_value - entry_value
        option_pnl += pnl
        greeks = core.strategy.calc_position_greeks(
            call_row,
            put_row,
            position["call_qty"],
            position["put_qty"],
            side=side,
        )
        greeks_list.append(greeks)
        signed_values.append(signed_value)
        margins.append(core.position.margin_value(position))
        rows.extend(
            [
                _account_leg_row(
                    report_date,
                    account_id,
                    side,
                    position,
                    call_row,
                    "call",
                    greeks,
                ),
                _account_leg_row(
                    report_date,
                    account_id,
                    side,
                    position,
                    put_row,
                    "put",
                    greeks,
                ),
            ]
        )
    return (
        rows,
        core.backtester.combine_greeks(greeks_list),
        sum(signed_values),
        sum(margins),
        option_pnl,
    )


def _account_leg_row(report_date, account_id, side, position, row, leg, greeks):
    qty_key = "call_qty" if leg == "call" else "put_qty"
    price_key = "entry_call_price" if leg == "call" else "entry_put_price"
    iv_key = "call_iv" if leg == "call" else "put_iv"
    delta_key = "call_delta" if leg == "call" else "put_delta"
    gamma_key = "call_gamma" if leg == "call" else "put_gamma"
    vega_key = "call_vega" if leg == "call" else "put_vega"
    theta_key = "call_theta" if leg == "call" else "put_theta"
    multiplier = _number(position.get("contract_multiplier")) or 0.0
    underlying_price = _first_number(
        row.get("underlying_close"),
        row.get("underlying_price"),
        row.get("underlying_last"),
    )
    mark_price = _option_mark_from_chain_row(row)
    return {
        "日期": report_date,
        "账户ID": account_id,
        "方向": side,
        "合约代码": row.get("order_book_id"),
        "合约名称": row.get("contract_symbol"),
        "买卖": "卖" if side == "short" else "买",
        "持仓类型": "义务仓" if side == "short" else "权利仓",
        "总持仓": position.get(qty_key),
        "今持仓": None,
        "今开仓": None,
        "今平仓": None,
        "可平量": None,
        "最新价": mark_price,
        "持仓均价": position.get(price_key),
        "开仓均价": position.get(price_key),
        "期权市值": (
            (mark_price or 0.0)
            * position.get(qty_key)
            * position.get("contract_multiplier")
        ),
        "占用保证金": position.get("option_margin") if leg == "call" else None,
        "持仓盈亏": None,
        "浮动盈亏": None,
        "AUM": _position_aum(position, multiplier, underlying_price),
        "行权价": row.get("strike_price"),
        "到期日": str(pd.Timestamp(row.get("maturity_date")).date()),
        "剩余天数": row.get("dte"),
        "IV": greeks.get(iv_key),
        "单张Delta": _single_option_delta_from_chain(row, side),
        "Delta": greeks.get(delta_key),
        "Gamma": greeks.get(gamma_key),
        "Vega": greeks.get(vega_key),
        "Theta": greeks.get(theta_key),
    }


def _position_aum(position, multiplier, underlying_price):
    if underlying_price is None or multiplier is None:
        return None
    call_qty = abs(_number(position.get("call_qty")) or 0.0)
    put_qty = abs(_number(position.get("put_qty")) or 0.0)
    return max(call_qty, put_qty) * float(multiplier) * float(underlying_price)


def _first_number(*values):
    for value in values:
        parsed = _number(value)
        if parsed is not None:
            return parsed
    return None


def _single_option_delta_from_chain(row, side):
    delta = _number(row.get("delta"))
    if delta is None:
        return None
    direction = -1.0 if str(side or "").lower() == "short" else 1.0
    return direction * delta


def _hedge_for_report_date(live_account, product, account_id, report_date):
    report_ts = pd.Timestamp(report_date).normalize()
    fills = account_store.list_fills(
        product,
        account_id=account_id,
        include_voided=False,
    )
    hedge = None
    has_hedge_fill = False
    saw_hedge_fill = False
    for row in fills:
        payload = _supported_fill(row["payload"])
        if payload is None:
            continue
        action = payload.get("action")
        if action not in {
            "delta_hedge",
            "rebalance_hedge",
            "close_hedge",
        }:
            continue
        has_hedge_fill = True
        fill_date = _date_or_none(payload.get("date"))
        if fill_date is None or fill_date > report_ts:
            continue
        saw_hedge_fill = True
        hedge = account_store.HedgeState(
            qty=float(payload.get("qty", payload.get("new_etf_qty", payload.get("target_hedge_qty", 0.0))) or 0.0),
            entry_price=float(payload.get("entry_price", payload.get("price", 0.0)) or 0.0),
            margin=float(payload.get("margin", 0.0) or 0.0),
            underlying_order_book_id=payload.get("underlying_order_book_id"),
        )
    if saw_hedge_fill:
        return hedge or account_store.HedgeState()
    if has_hedge_fill:
        return account_store.HedgeState()
    return live_account.hedge


def _hedge_rows_from_account(
    product,
    hedge,
    account_id,
    report_date,
    spot,
    prefer_spot_mark=False,
    not_before=None,
):
    if abs(float(hedge.qty or 0.0)) <= 1e-9:
        return []

    mark_with_spot = _can_mark_hedge_with_spot(product, hedge)
    if mark_with_spot:
        latest_price = float(spot)
    else:
        latest_price = hedge.latest_price
    if latest_price is None:
        latest_price = float(spot)
    qty = float(hedge.qty)
    market_value = None if mark_with_spot else hedge.last_market_value
    if market_value is None:
        market_value = qty * latest_price
    entry_price = _hedge_open_cost_for_report(
        product,
        account_id,
        report_date,
    )
    if entry_price is None:
        entry_price = hedge.entry_price
    floating_pnl = None
    if (_number(entry_price) or 0.0) > 0:
        floating_pnl = core.hedge.calc_unrealized_pnl(qty, entry_price, latest_price)
    if floating_pnl is None:
        floating_pnl = 0.0
    security_code = _security_code_from_underlying(hedge.underlying_order_book_id)
    security_name = hedge.underlying_order_book_id
    return [
        {
            "日期": report_date,
            "账户ID": account_id,
            "方向": "hedge",
            "合约代码": security_code or hedge.underlying_order_book_id,
            "合约名称": security_name,
            "买卖": "买" if qty > 0 else "卖",
            "持仓类型": "ETF对冲",
            "总持仓": qty,
            "今持仓": None,
            "今开仓": None,
            "今平仓": None,
            "可平量": None,
            "最新价": latest_price,
            "持仓均价": entry_price,
            "开仓均价": entry_price,
            "期权市值": market_value,
            "占用保证金": hedge.margin,
            "持仓盈亏": floating_pnl,
            "浮动盈亏": floating_pnl,
            "行权价": None,
            "到期日": None,
            "剩余天数": None,
            "IV": None,
            "单张Delta": None,
            "Delta": qty,
            "Gamma": 0.0,
            "Vega": 0.0,
            "Theta": 0.0,
        }
    ]


def _can_mark_hedge_with_spot(product, hedge):
    spec = market_data.SSE_ETF_OPTION_SPECS.get(product)
    if spec is None:
        return False
    underlying = str(hedge.underlying_order_book_id or "")
    if not underlying:
        return True
    return underlying in {
        spec.etf_symbol,
        spec.etf_file_prefix,
        f"sh{spec.etf_symbol}",
    }


def _trade_rows_from_export(product, report_date, not_before=None):
    path = _latest_export_file("成交明细", report_date, not_before=not_before)
    if path is None:
        return []
    rows = _trade_rows_from_file(path, product)
    return [row for row in rows if _date8_to_iso(row.get("日期")) == report_date]


def _etf_trade_rows_from_export(product, report_date, not_before=None):
    path = _latest_export_file(
        "证券委托查询_实时成交(信息导出)",
        report_date,
        not_before=not_before,
    )
    if path is None:
        return []
    return _etf_trade_rows_from_file(path, product, report_date)


def _etf_trade_rows_from_file(path, product, report_date):
    target_code = market_data.SSE_ETF_OPTION_SPECS[product].etf_symbol
    rows = []
    for _, item in _read_export_csv(path).iterrows():
        code = _security_code(item.get("证券代码"))
        if code != target_code or _date8_to_iso(item.get("日期")) != report_date:
            continue
        price = _number(item.get("成交价格"))
        qty = _number(item.get("成交数量"))
        if price is None or qty is None or qty <= 0:
            continue
        rows.append(
            {
                "序号": item.get("序号"),
                "投资者账号": item.get("投资者账号"),
                "交易所": item.get("交易所"),
                "合约代码": code,
                "合约名称": item.get("证券名称"),
                "成交编号": item.get("成交编号") or item.get("报单编号"),
                "报单编号": item.get("报单编号"),
                "开平": None,
                "买卖": _clean_text(item.get("买卖")),
                "报单价格": _number(item.get("报单价格")),
                "成交价格": price,
                "成交数量": qty,
                "手续费": _configured_etf_trade_fee(product, price, qty),
                "平仓盈亏": None,
                "类型": "ETF对冲",
                "日期": report_date,
                "报单时间": item.get("报单时间"),
                "成交时间": item.get("成交时间"),
                "成交时间(日)": item.get("成交时间(日)"),
                "策略名称": item.get("策略名称"),
            }
        )
    return rows


def _all_etf_trade_rows_from_exports(product, not_before=None):
    rows = []
    for path in sorted(
        _live_hold_dir().glob("证券委托查询_实时成交(信息导出)*.csv")
    ):
        if not _export_file_is_not_before(path, not_before):
            continue
        report_date = _filename_date(path)
        rows.extend(_etf_trade_rows_from_file(path, product, report_date))
    seen = set()
    unique = []
    for row in rows:
        key = row.get("成交编号") or (row.get("合约代码"), row.get("成交时间(日)"))
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def _all_trade_rows_from_exports(product, not_before=None):
    rows = []
    for path in sorted(_live_hold_dir().glob("成交明细*.csv")):
        if not _export_file_is_not_before(path, not_before):
            continue
        rows.extend(_trade_rows_from_file(path, product))
    seen = set()
    unique = []
    for row in rows:
        key = row.get("成交编号") or (row.get("合约代码"), row.get("成交时间(日)"))
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return sorted(unique, key=lambda row: str(row.get("成交时间(日)") or ""))


def _greeks_trade_rows_from_exports(product, account_id="default"):
    if product is None:
        return []
    not_before = None
    try:
        not_before = account_store.load_account(product, account_id=account_id).reset_at
    except Exception:
        pass
    rows = _all_trade_rows_from_exports(product, not_before=not_before)
    rows.extend(_all_etf_trade_rows_from_exports(product, not_before=not_before))
    return rows


def _add_summary_greeks_pnl(
    summary_history,
    position_history=None,
    product=None,
    current_position_report=None,
    trade_rows=None,
    account_id="default",
):
    history = summary_history.copy()
    for column in SUMMARY_COLUMNS:
        if column not in history.columns:
            history[column] = None
    if history.empty:
        return history.reindex(columns=SUMMARY_COLUMNS)

    history = _fill_summary_leg_greeks_from_positions(
        history,
        position_history,
        product,
    )
    history = _fill_summary_hedge_marks_and_pnl(
        history,
        position_history,
        product,
    )
    if trade_rows is None:
        trade_rows = _greeks_trade_rows_from_exports(product, account_id)
    history = history.sort_values(["账户ID", "日期"]).reset_index(drop=True)
    groups = []
    for _, group in history.groupby("账户ID", dropna=False, sort=False):
        groups.append(
            _add_summary_greeks_pnl_for_account(
                group.copy(),
                product,
                position_history=position_history,
                trade_rows=trade_rows,
            )
        )
    result = pd.concat(groups, ignore_index=True) if groups else history
    unavailable = result["GreeksPnL说明"].astype(str).eq("first_history_row")
    result.loc[unavailable, DAILY_GREEKS_PNL_COLUMNS] = np.nan
    result.loc[unavailable, "GreeksPnL口径"] = "unavailable"
    result = _override_vega_pnl_with_same_contract_iv(
        result,
        position_history,
        current_position_report,
        product,
    )
    result = _refresh_summary_pnl_reconciliation_residual(result)
    result = result.sort_values(["日期", "账户ID"]).reset_index(drop=True)
    return result.reindex(columns=SUMMARY_COLUMNS)


def _override_vega_pnl_with_same_contract_iv(
    summary_history,
    position_history,
    current_position_report,
    product=None,
):
    if (
        position_history is None
        or position_history.empty
        or summary_history.empty
    ):
        return summary_history
    required_history = {"日期", "账户ID", "方向", "合约代码", "IV", "Vega"}
    if not required_history.issubset(position_history.columns):
        return summary_history
    required_report = {"日期", "合约代码", "IV"}

    result = summary_history.copy()
    positions = position_history.copy()
    positions = positions.loc[~positions["方向"].astype(str).eq("hedge")].copy()
    if (
        current_position_report is not None
        and not current_position_report.empty
        and required_report.issubset(current_position_report.columns)
    ):
        report = current_position_report.copy()
        report = report.loc[report["到期日"].notna()] if "到期日" in report.columns else report
        report["_code"] = report["合约代码"].apply(_security_code)
        report["_date"] = report["日期"].astype(str)
    else:
        report = pd.DataFrame()

    result = result.sort_values(["账户ID", "日期"]).copy()
    for _, indexes in result.groupby("账户ID", dropna=False, sort=False).groups.items():
        ordered = list(indexes)
        for offset, current_index in enumerate(ordered):
            if offset == 0:
                continue
            account_id = str(result.at[current_index, "账户ID"])
            current_date = str(result.at[current_index, "日期"])
            previous_date = str(result.at[ordered[offset - 1], "日期"])
            previous_positions = positions.loc[
                positions["账户ID"].astype(str).eq(account_id)
                & positions["日期"].astype(str).eq(previous_date)
            ]
            current_positions = positions.loc[
                positions["账户ID"].astype(str).eq(account_id)
                & positions["日期"].astype(str).eq(current_date)
            ]
            current_report = (
                report.loc[report["_date"].eq(current_date)]
                if not report.empty
                else pd.DataFrame()
            )
            current_spot = _number(result.at[current_index, "标的价格"])
            if previous_positions.empty:
                continue

            vega_pnl = 0.0
            matched = False
            for _, previous in previous_positions.iterrows():
                code = _security_code(previous.get("合约代码"))
                if code is None:
                    continue
                previous_vega = _number(previous.get("Vega"))
                previous_iv = _number(previous.get("IV"))
                current_iv = _current_option_iv_for_previous_position(
                    product,
                    current_date,
                    code,
                    previous,
                    current_spot,
                    current_report,
                    current_positions,
                )
                if previous_vega is None or previous_iv is None or current_iv is None:
                    continue
                vega_pnl += previous_vega * (current_iv - previous_iv) * 100.0
                matched = True

            if not matched:
                continue
            result.at[current_index, "期权单日VegaPnL"] = vega_pnl
            result.at[current_index, "单日VegaPnL"] = vega_pnl
            option_greeks = sum(
                _number(result.at[current_index, column]) or 0.0
                for column in [
                    "期权单日DeltaPnL",
                    "期权单日GammaPnL",
                    "期权单日VegaPnL",
                    "期权单日ThetaPnL",
                ]
            )
            hedge_greeks = _number(result.at[current_index, "对冲单日GreeksPnL"]) or 0.0
            transaction_greeks = _number(result.at[current_index, "交易GreeksPnL"]) or 0.0
            result.at[current_index, "期权单日GreeksPnL"] = option_greeks
            result.at[current_index, "昨仓GreeksPnL"] = option_greeks + hedge_greeks
            result.at[current_index, "单日VegaPnL"] = (
                vega_pnl + (_number(result.at[current_index, "交易VegaPnL"]) or 0.0)
            )
            result.at[current_index, "单日GreeksPnL"] = (
                option_greeks + hedge_greeks + transaction_greeks
            )
            result.at[current_index, "GreeksPnL说明"] = (
                "previous_close_same_contract_iv_for_vega_plus_intraday_trades"
                if abs(transaction_greeks) > 1e-9
                else "previous_close_same_contract_iv_for_vega"
            )
    return result


def _refresh_summary_pnl_reconciliation_residual(summary_history):
    if summary_history is None or summary_history.empty:
        return summary_history
    required = {"总单日盈亏", "当日盈亏分解合计", "当日盈亏对账差额"}
    if not required.issubset(summary_history.columns):
        return summary_history
    result = summary_history.copy()
    total = pd.to_numeric(result["总单日盈亏"], errors="coerce")
    decomposition = pd.to_numeric(result["当日盈亏分解合计"], errors="coerce")
    valid = total.notna() & decomposition.notna()
    result.loc[valid, "当日盈亏对账差额"] = total.loc[valid] - decomposition.loc[valid]
    return result


def _close_snapshot_option_daily_pnl(
    product,
    report_date,
    previous_positions,
    current_positions,
    trade_rows,
):
    if previous_positions.empty and current_positions.empty:
        return None

    multiplier = _contract_multiplier(product)
    previous_quantities = _signed_option_quantities(previous_positions)
    remaining_previous_quantities = dict(previous_quantities)
    previous_prices = {}
    for _, row in previous_positions.iterrows():
        code = _security_code(row.get("合约代码"))
        price = _number(row.get("最新价"))
        if code is not None and price is not None:
            previous_prices[code] = price

    current_marks = {}
    for _, row in current_positions.iterrows():
        code = _security_code(row.get("合约代码"))
        price = _number(row.get("最新价"))
        if code is not None and price is not None:
            current_marks[code] = price

    option_trades = [
        row
        for row in trade_rows
        if _security_code(row.get("合约代码")) is not None
        and str(_security_code(row.get("合约代码"))).startswith("100")
        and str(row.get("类型") or "").upper().find("ETF") < 0
    ]
    option_trades = sorted(
        option_trades,
        key=lambda row: (
            _trade_row_timestamp(row, report_date) or pd.Timestamp.max,
            str(row.get("成交编号") or row.get("报单编号") or ""),
        ),
    )

    close_trade_pnl = 0.0
    for trade in option_trades:
        code = _security_code(trade.get("合约代码"))
        price = _number(trade.get("成交价格"))
        if code is None or price is None:
            return None
        signed_trade_qty = _signed_trade_quantity(trade)
        remaining_qty = remaining_previous_quantities.get(code, 0.0)
        is_close = str(trade.get("开平") or "").find("平") >= 0
        if is_close:
            realized = _number(trade.get("平仓盈亏"))
            if realized is not None:
                close_trade_pnl += realized
            else:
                closed_qty = _closed_signed_quantity(remaining_qty, signed_trade_qty)
                previous_price = previous_prices.get(code)
                if abs(closed_qty) > 1e-9:
                    if previous_price is None:
                        return None
                    close_trade_pnl += (
                        closed_qty * (float(price) - float(previous_price)) * multiplier
                    )
        closed_qty = _closed_signed_quantity(remaining_qty, signed_trade_qty)
        if abs(closed_qty) > 1e-9:
            remaining_previous_quantities[code] = remaining_qty - closed_qty

    holding_pnl = 0.0
    for code, qty in remaining_previous_quantities.items():
        if abs(qty) <= 1e-9:
            continue
        previous_price = previous_prices.get(code)
        current_price = current_marks.get(code)
        if previous_price is None or current_price is None:
            return None
        holding_pnl += qty * (float(current_price) - float(previous_price)) * multiplier

    return float(holding_pnl + close_trade_pnl)


def _closed_signed_quantity(remaining_previous_qty, signed_trade_qty):
    if abs(remaining_previous_qty) <= 1e-9 or abs(signed_trade_qty) <= 1e-9:
        return 0.0
    if remaining_previous_qty * signed_trade_qty >= 0:
        return 0.0
    magnitude = min(abs(remaining_previous_qty), abs(signed_trade_qty))
    return math.copysign(magnitude, remaining_previous_qty)


def _option_position_rows_by_code(
    previous_positions,
    current_positions,
    previous_date=None,
    current_date=None,
):
    rows = {}
    for label, frame in [("previous", previous_positions), ("current", current_positions)]:
        if frame is None or frame.empty:
            continue
        for _, row in frame.iterrows():
            code = _security_code(row.get("合约代码"))
            if code is None:
                continue
            rows.setdefault(code, {})[label] = row
    for values in rows.values():
        values["_previous_date"] = previous_date
        values["_current_date"] = current_date
    return rows


def _signed_option_quantities(position_rows):
    quantities = {}
    if position_rows is None or position_rows.empty:
        return quantities
    for _, row in position_rows.iterrows():
        code = _security_code(row.get("合约代码"))
        if code is None:
            continue
        qty = abs(_number(row.get("总持仓")) or 0.0)
        direction = -1.0 if str(row.get("方向") or "").lower() == "short" else 1.0
        quantities[code] = quantities.get(code, 0.0) + direction * qty
    return quantities


def _signed_trade_quantity(trade):
    qty = abs(_number(trade.get("成交数量")) or 0.0)
    if qty <= 0:
        return 0.0
    direction = str(trade.get("买卖") or "")
    return -qty if "卖" in direction else qty


def _option_price_from_quote_snapshot(product, report_date, code, timestamp):
    root = (
        storage.PROJECT_ROOT
        / "data"
        / "live"
        / product
        / "quotes"
        / pd.Timestamp(report_date).strftime("%Y%m%d")
    )
    if not root.exists():
        return None
    candidates = []
    for path in root.glob("*_option_chain.parquet"):
        prefix = path.name.split("_", 1)[0]
        if not re.fullmatch(r"\d{6}", prefix):
            continue
        snapshot_ts = pd.Timestamp(
            f"{pd.Timestamp(report_date).date()} "
            f"{prefix[:2]}:{prefix[2:4]}:{prefix[4:6]}"
        )
        if snapshot_ts <= timestamp:
            candidates.append((snapshot_ts, path))
    if not candidates:
        return None
    _, path = sorted(candidates, key=lambda item: item[0])[-1]
    try:
        frame = pd.read_parquet(path)
    except Exception:
        return None
    if frame.empty or "order_book_id" not in frame.columns:
        return None
    rows = frame.loc[frame["order_book_id"].apply(_security_code).eq(code)]
    if rows.empty:
        return None
    row = rows.iloc[0]
    for column in ["close", "last", "latest", "price", "mid"]:
        value = _number(row.get(column))
        if value is not None and value > 0:
            return value
    bid = _number(row.get("bid"))
    ask = _number(row.get("ask"))
    if bid is not None and ask is not None and bid > 0 and ask > 0:
        return (bid + ask) / 2.0
    return bid if bid is not None and bid > 0 else ask


def _option_greeks_for_dte(
    product,
    row,
    price,
    spot,
    flag,
    signed_qty,
    dte,
    fallback_iv=None,
):
    strike = _number(row.get("行权价"))
    if strike is None or strike <= 0:
        return None
    if dte is None or dte <= 0:
        return None
    config = load_product_config(product)
    ttm = float(dte) / float(config.vol.annual_days)
    vollib = core.vol_engine._load_vollib_funcs()
    try:
        iv = vollib["implied_volatility"](
            price=pd.Series([float(price)]),
            S=pd.Series([float(spot)]),
            t=pd.Series([ttm]),
            K=float(strike),
            r=float(config.vol.risk_free_rate),
            flag=flag,
            model="black_scholes",
            return_as="series",
            on_error="ignore",
        ).iloc[0]
    except Exception:
        iv = 0.0
    iv = _number(iv)
    if iv is None or iv < 0:
        iv = 0.0
    if iv <= 0:
        iv = fallback_iv
    if iv is None or iv <= 0:
        return None
    kwargs = {
        "flag": flag,
        "S": pd.Series([float(spot)]),
        "K": float(strike),
        "t": pd.Series([ttm]),
        "r": float(config.vol.risk_free_rate),
        "model": "black_scholes",
        "sigma": pd.Series([float(iv)]),
        "return_as": "series",
    }
    try:
        delta = float(vollib["delta"](**kwargs).iloc[0])
        gamma = float(vollib["gamma"](**kwargs).iloc[0])
        vega = float(vollib["vega"](**kwargs).iloc[0])
        theta_365 = float(vollib["theta"](**kwargs).iloc[0])
    except Exception:
        return None
    scale = float(signed_qty) * _contract_multiplier(product)
    theta = theta_365 * (365.0 / float(config.vol.annual_days))
    return {
        "iv": float(iv),
        "delta": delta * scale,
        "gamma": gamma * scale,
        "vega": vega * scale,
        "theta": theta * scale,
    }


def _trade_row_timestamp(row, report_date):
    value = (
        row.get("成交时间(日)")
        or row.get("成交时间")
        or row.get("报单时间")
    )
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", text):
        try:
            return pd.Timestamp(f"{report_date} {text}")
        except Exception:
            return None
    try:
        timestamp = pd.Timestamp(text)
    except Exception:
        try:
            timestamp = pd.Timestamp(f"{report_date} {text}")
        except Exception:
            return None
    if timestamp.date() == pd.Timestamp(report_date).date():
        return timestamp
    return pd.Timestamp.combine(pd.Timestamp(report_date).date(), timestamp.time())


def _spot_from_intraday_minute(product, report_date, timestamp):
    detail = _spot_from_intraday_minute_detail(product, report_date, timestamp)
    if detail is None:
        return None
    return detail["price"]


def _spot_from_intraday_minute_detail(product, report_date, timestamp):
    spec = market_data.SSE_ETF_OPTION_SPECS.get(product)
    if spec is None:
        return None
    path = (
        storage.PROJECT_ROOT
        / "data"
        / "live"
        / product
        / "intraday"
        / pd.Timestamp(report_date).strftime("%Y%m%d")
        / f"etf_{spec.etf_symbol}_1m.csv"
    )
    if not path.exists():
        return None
    try:
        frame = pd.read_csv(path, encoding="utf-8-sig")
    except Exception:
        return None
    if "timestamp" not in frame.columns or "close" not in frame.columns:
        return None
    frame["timestamp"] = pd.to_datetime(frame["timestamp"], errors="coerce")
    frame["close"] = pd.to_numeric(frame["close"], errors="coerce")
    frame = frame.dropna(subset=["timestamp", "close"])
    frame = frame.loc[frame["timestamp"].le(timestamp)].sort_values("timestamp")
    if frame.empty:
        return None
    latest = frame.iloc[-1]
    latest_timestamp = pd.Timestamp(latest["timestamp"])
    if pd.Timestamp(timestamp) - latest_timestamp > pd.Timedelta(minutes=5):
        return None
    return {
        "source": "intraday_minute",
        "timestamp": latest_timestamp,
        "price": float(latest["close"]),
        "path": str(path),
    }


def _spot_from_quote_snapshot(product, report_date, timestamp):
    detail = _spot_from_quote_snapshot_detail(product, report_date, timestamp)
    if detail is None:
        return None
    return detail["price"]


def _spot_from_quote_snapshot_detail(product, report_date, timestamp):
    root = (
        storage.PROJECT_ROOT
        / "data"
        / "live"
        / product
        / "quotes"
        / pd.Timestamp(report_date).strftime("%Y%m%d")
    )
    if not root.exists():
        return None
    candidates = []
    for path in root.glob("*_etf.parquet"):
        prefix = path.name.split("_", 1)[0]
        if not re.fullmatch(r"\d{6}", prefix):
            continue
        snapshot_ts = pd.Timestamp(
            f"{pd.Timestamp(report_date).date()} "
            f"{prefix[:2]}:{prefix[2:4]}:{prefix[4:6]}"
        )
        if snapshot_ts <= timestamp:
            candidates.append((snapshot_ts, path))
    if not candidates:
        return None
    snapshot_ts, path = sorted(candidates, key=lambda item: item[0])[-1]
    try:
        frame = pd.read_parquet(path)
    except Exception:
        return None
    if frame.empty or "close" not in frame.columns:
        return None
    close = _number(frame.iloc[-1].get("close"))
    if close is None:
        return None
    return {
        "source": "quote_snapshot",
        "timestamp": snapshot_ts,
        "price": close,
        "path": str(path),
    }


def _add_summary_greeks_pnl_for_account(
    group,
    product=None,
    position_history=None,
    trade_rows=None,
):
    spot = _numeric_series(group, "标的价格")
    existing_option_daily_pnl = _numeric_series(group, "期权单日盈亏")
    existing_etf_daily_pnl = _numeric_series(group, "ETF单日盈亏")
    legacy_hedge_daily_pnl = _numeric_series(group, "对冲单日盈亏")
    existing_etf_daily_pnl = existing_etf_daily_pnl.combine_first(
        legacy_hedge_daily_pnl
    )
    raw_option_actual_pnl = _actual_daily_pnl_series_from_positions(
        group,
        position_history,
        product,
        side="option",
    )
    raw_option_actual_pnl = raw_option_actual_pnl.combine_first(
        _legacy_total_daily_pnl_series(group, "期权总盈亏", "期权浮盈亏")
    )
    option_actual_pnl = _previous_close_option_pnl_series(
        raw_option_actual_pnl,
        group,
        position_history,
        product,
    ).combine_first(raw_option_actual_pnl)
    raw_etf_actual_pnl = _actual_daily_pnl_series_from_positions(
        group,
        position_history,
        product,
        side="hedge",
    )
    raw_etf_actual_pnl = raw_etf_actual_pnl.combine_first(
        _legacy_total_daily_pnl_series(group, "对冲总盈亏", "对冲浮盈亏")
    )
    hedge_mark = _numeric_series(group, "对冲最新价")
    hedge_qty = _numeric_series(group, "对冲持仓").fillna(0.0)
    etf_actual_pnl = _previous_close_hedge_pnl_series(
        raw_etf_actual_pnl,
        spot,
        hedge_mark,
        hedge_qty,
    )
    call_iv = _numeric_series(group, "Call IV")
    put_iv = _numeric_series(group, "Put IV")
    call_delta = _numeric_series(group, "Call Delta")
    put_delta = _numeric_series(group, "Put Delta")
    call_gamma = _numeric_series(group, "Call Gamma")
    put_gamma = _numeric_series(group, "Put Gamma")
    call_vega = _numeric_series(group, "Call Vega")
    put_vega = _numeric_series(group, "Put Vega")
    call_theta = _numeric_series(group, "Call Theta")
    put_theta = _numeric_series(group, "Put Theta")

    spot_chg = spot.diff()
    option_explainable = (
        spot.shift(1).notna()
        & call_iv.shift(1).notna()
        & put_iv.shift(1).notna()
        & call_iv.notna()
        & put_iv.notna()
    )

    option_delta_pnl = (
        call_delta.shift(1)
        + put_delta.shift(1)
    ) * spot_chg
    hedge_delta_pnl = _previous_close_hedge_pnl_series(
        raw_etf_actual_pnl,
        spot,
        hedge_mark,
        hedge_qty,
    )
    gamma_pnl = (
        0.5
        * (
            call_gamma.shift(1)
            + put_gamma.shift(1)
        )
        * spot_chg**2
    )
    vega_pnl = (
        call_vega.shift(1) * (call_iv - call_iv.shift(1)) * 100
        + put_vega.shift(1) * (put_iv - put_iv.shift(1)) * 100
    )
    theta_pnl = (
        call_theta.shift(1)
        + put_theta.shift(1)
    ) * _trading_day_steps(group["日期"], product=product)
    transaction_greeks = _transaction_greeks_pnl_series(
        product,
        group,
        position_history,
        trade_rows,
    )

    group["期权单日DeltaPnL"] = option_delta_pnl.where(option_explainable, 0.0).fillna(0.0)
    group["期权单日GammaPnL"] = gamma_pnl.where(option_explainable, 0.0).fillna(0.0)
    group["期权单日VegaPnL"] = vega_pnl.where(option_explainable, 0.0).fillna(0.0)
    group["期权单日ThetaPnL"] = theta_pnl.where(option_explainable, 0.0).fillna(0.0)
    group["期权单日GreeksPnL"] = group[
        ["期权单日DeltaPnL", "期权单日GammaPnL", "期权单日VegaPnL", "期权单日ThetaPnL"]
    ].sum(axis=1)
    group["对冲单日DeltaPnL"] = hedge_delta_pnl.fillna(0.0)
    group["对冲单日GreeksPnL"] = group["对冲单日DeltaPnL"]
    for source, target in [
        ("delta_pnl", "交易DeltaPnL"),
        ("gamma_pnl", "交易GammaPnL"),
        ("vega_pnl", "交易VegaPnL"),
        ("theta_pnl", "交易ThetaPnL"),
        ("greeks_pnl", "交易GreeksPnL"),
    ]:
        group[target] = transaction_greeks[source]
    group["昨仓GreeksPnL"] = group["期权单日GreeksPnL"] + group["对冲单日GreeksPnL"]
    group["单日DeltaPnL"] = (
        group["期权单日DeltaPnL"]
        + group["对冲单日DeltaPnL"]
        + group["交易DeltaPnL"]
    )
    group["单日GammaPnL"] = group["期权单日GammaPnL"] + group["交易GammaPnL"]
    group["单日VegaPnL"] = group["期权单日VegaPnL"] + group["交易VegaPnL"]
    group["单日ThetaPnL"] = group["期权单日ThetaPnL"] + group["交易ThetaPnL"]
    group["单日GreeksPnL"] = group[
        ["单日DeltaPnL", "单日GammaPnL", "单日VegaPnL", "单日ThetaPnL"]
    ].sum(axis=1)
    group["期权单日盈亏"] = _prefer_existing_daily_pnl(
        existing_option_daily_pnl,
        option_actual_pnl,
    ).fillna(0.0)
    group["ETF单日盈亏"] = _prefer_existing_daily_pnl(
        existing_etf_daily_pnl,
        etf_actual_pnl,
    ).fillna(0.0)
    if "对冲单日盈亏" in group.columns:
        group["对冲单日盈亏"] = group["ETF单日盈亏"]
    group["总单日盈亏"] = group[["期权单日盈亏", "ETF单日盈亏"]].sum(axis=1)
    group["券商期权单日盈亏变化"] = group["期权单日盈亏"]
    group["券商对冲单日盈亏变化"] = group["ETF单日盈亏"]
    group["券商总单日盈亏变化"] = group["总单日盈亏"]
    has_transaction_greeks = group["交易GreeksPnL"].abs().gt(1e-9)
    group["GreeksPnL口径"] = "previous_close"
    group["GreeksPnL说明"] = "all_greeks_use_previous_close"
    group.loc[has_transaction_greeks, "GreeksPnL口径"] = (
        "previous_close_plus_transaction_to_close"
    )
    group.loc[has_transaction_greeks, "GreeksPnL说明"] = (
        "previous_close_positions_plus_intraday_trades"
    )
    group["GreeksPnL路径节点数"] = None
    if not group.empty:
        group.iloc[0, group.columns.get_loc("GreeksPnL说明")] = "first_history_row"

    return group


def _prefer_existing_daily_pnl(existing, computed):
    result = existing.combine_first(computed)
    if result.empty:
        return result
    first_index = result.index[0]
    existing_first = _number(existing.loc[first_index]) if first_index in existing.index else None
    computed_first = _number(computed.loc[first_index]) if first_index in computed.index else None
    if (
        existing_first is not None
        and abs(existing_first) <= 1e-9
        and computed_first is not None
        and abs(computed_first) > 1e-9
    ):
        result.loc[first_index] = computed_first
    return result


def _transaction_greeks_pnl_series(product, group, position_history=None, trade_rows=None):
    columns = ["delta_pnl", "gamma_pnl", "vega_pnl", "theta_pnl", "greeks_pnl"]
    result = pd.DataFrame(0.0, index=group.index, columns=columns)
    if product is None or group.empty or not trade_rows:
        return result
    trades_by_date = _trade_rows_by_date(trade_rows)
    ordered = group.sort_values("日期")
    for offset, (index, current) in enumerate(ordered.iterrows()):
        current_date = str(pd.Timestamp(current.get("日期")).date())
        day_trades = trades_by_date.get(current_date, [])
        if not day_trades:
            continue
        account_id = str(current.get("账户ID") or "default")
        previous_date = (
            str(pd.Timestamp(ordered.iloc[offset - 1].get("日期")).date())
            if offset > 0
            else None
        )
        current_positions = _positions_for_summary_date(
            position_history,
            current_date,
            account_id,
        )
        previous_positions = _positions_for_summary_date(
            position_history,
            previous_date,
            account_id,
        )
        parts = _transaction_greeks_pnl_for_day(
            product,
            current_date,
            current,
            previous_positions,
            current_positions,
            day_trades,
        )
        for column in columns:
            result.at[index, column] = parts[column]
    result["greeks_pnl"] = result[
        ["delta_pnl", "gamma_pnl", "vega_pnl", "theta_pnl"]
    ].sum(axis=1)
    return result


def _trade_rows_by_date(trade_rows):
    result = {}
    for row in trade_rows or []:
        date = _date_or_none(row.get("日期"))
        if date is None:
            timestamp = _trade_row_timestamp(row, pd.Timestamp.today().date())
            date = timestamp if timestamp is not None else None
        if date is None:
            continue
        result.setdefault(str(pd.Timestamp(date).date()), []).append(row)
    return result


def _positions_for_summary_date(position_history, date, account_id):
    if position_history is None or position_history.empty or date is None:
        return pd.DataFrame()
    required = {"日期", "账户ID"}
    if not required.issubset(position_history.columns):
        return pd.DataFrame()
    return position_history.loc[
        position_history["日期"].astype(str).eq(str(date))
        & position_history["账户ID"].astype(str).eq(str(account_id))
    ].copy()


def _transaction_greeks_pnl_for_day(
    product,
    report_date,
    summary_row,
    previous_positions,
    current_positions,
    trade_rows,
):
    parts = {
        "delta_pnl": 0.0,
        "gamma_pnl": 0.0,
        "vega_pnl": 0.0,
        "theta_pnl": 0.0,
        "greeks_pnl": 0.0,
    }
    close_spot = _number(summary_row.get("标的价格"))
    hedge_close = _number(summary_row.get("对冲最新价"))
    if hedge_close is None:
        hedge_close = close_spot
    rows_by_code = _option_position_rows_by_code(previous_positions, current_positions)
    for trade in trade_rows or []:
        if str(trade.get("类型") or "") == "ETF对冲":
            delta = _transaction_hedge_delta_pnl(trade, hedge_close)
            if delta is not None:
                parts["delta_pnl"] += delta
            continue
        contribution = _transaction_option_greeks_pnl(
            product,
            report_date,
            trade,
            rows_by_code,
            close_spot,
        )
        if contribution is None:
            continue
        for key in ["delta_pnl", "gamma_pnl", "vega_pnl", "theta_pnl"]:
            parts[key] += contribution[key]
    parts["greeks_pnl"] = sum(parts[key] for key in ["delta_pnl", "gamma_pnl", "vega_pnl", "theta_pnl"])
    return parts


def _transaction_hedge_delta_pnl(trade, hedge_close):
    if hedge_close is None:
        return None
    trade_price = _number(trade.get("成交价格"))
    signed_qty = _signed_trade_quantity(trade)
    if trade_price is None or abs(signed_qty) <= 1e-9:
        return None
    return float(signed_qty) * (float(hedge_close) - float(trade_price))


def _transaction_option_greeks_pnl(product, report_date, trade, rows_by_code, close_spot):
    if close_spot is None:
        return None
    code = _security_code(trade.get("合约代码"))
    trade_price = _number(trade.get("成交价格"))
    signed_qty = _signed_trade_quantity(trade)
    if code is None or trade_price is None or abs(signed_qty) <= 1e-9:
        return None
    rows = rows_by_code.get(code)
    if not rows:
        return None
    reference = rows.get("current")
    if reference is None:
        reference = rows.get("previous")
    if reference is None:
        return None
    leg = _position_row_leg(reference)
    flag = "c" if leg == "Call" else "p" if leg == "Put" else None
    if flag is None:
        return None
    timestamp = _trade_row_timestamp(trade, report_date)
    if timestamp is None:
        timestamp = pd.Timestamp(f"{report_date} 15:00:00")
    start_spot = _spot_from_intraday_minute(product, report_date, timestamp)
    if start_spot is None:
        start_spot = _spot_from_quote_snapshot(product, report_date, timestamp)
    if start_spot is None:
        start_spot = close_spot
    close_price = _transaction_option_close_price(product, report_date, code, rows)
    if close_price is None:
        return None
    close_dte = _option_close_dte(product, report_date, rows, reference)
    if close_dte is None or close_dte <= 0:
        return None
    start_dte = float(close_dte) + _remaining_trading_day_fraction(
        timestamp,
        report_date,
    )
    fallback_iv = _transaction_reference_iv(rows, reference)
    start_greeks = _option_greeks_for_dte(
        product,
        reference,
        float(trade_price),
        float(start_spot),
        flag,
        float(signed_qty),
        start_dte,
        fallback_iv=fallback_iv,
    )
    end_greeks = _option_greeks_for_dte(
        product,
        reference,
        float(close_price),
        float(close_spot),
        flag,
        float(signed_qty),
        float(close_dte),
        fallback_iv=fallback_iv,
    )
    if start_greeks is None or end_greeks is None:
        return None
    spot_change = float(close_spot) - float(start_spot)
    theta_fraction = _trade_to_close_fraction(timestamp, report_date)
    return {
        "delta_pnl": start_greeks["delta"] * spot_change,
        "gamma_pnl": 0.5 * start_greeks["gamma"] * spot_change * spot_change,
        "vega_pnl": start_greeks["vega"] * (end_greeks["iv"] - start_greeks["iv"]) * 100.0,
        "theta_pnl": start_greeks["theta"] * theta_fraction,
    }


def _option_close_dte(product, report_date, rows, reference):
    calendar = market_data.load_live_trading_calendar()
    for row in [
        rows.get("current") if rows else None,
        reference,
        rows.get("previous") if rows else None,
    ]:
        if row is None:
            continue
        maturity = row.get("到期日")
        if maturity is None or pd.isna(maturity):
            continue
        try:
            return float(
                core.vol_engine._count_trading_dte(
                    report_date,
                    maturity,
                    trading_calendar=calendar,
                )
            )
        except (TypeError, ValueError):
            continue

    current = rows.get("current") if rows else None
    current_dte = _number(current.get("剩余天数")) if current is not None else None
    if current_dte is not None and current_dte > 0:
        return float(current_dte)

    previous = rows.get("previous") if rows else None
    previous_dte = _number(previous.get("剩余天数")) if previous is not None else None
    if previous_dte is not None and previous_dte > 0:
        previous_date = previous.get("日期")
        try:
            elapsed = core.vol_engine._count_trading_dte(
                previous_date,
                report_date,
                trading_calendar=calendar,
            )
        except (TypeError, ValueError):
            elapsed = 1
        return max(float(previous_dte) - float(elapsed), 0.0)

    return None


def _transaction_reference_iv(rows, reference):
    for row in [
        rows.get("current") if rows else None,
        reference,
        rows.get("previous") if rows else None,
    ]:
        if row is None:
            continue
        iv = _number(row.get("IV"))
        if iv is not None and iv > 0:
            return float(iv)
    return None


def _transaction_option_close_price(product, report_date, code, rows):
    current = rows.get("current")
    if current is not None:
        price = _number(current.get("最新价"))
        if price is not None and price > 0:
            return price
    price = _option_close_price_from_quote_snapshot(product, report_date, code)
    if price is not None and price > 0:
        return price
    previous = rows.get("previous")
    if previous is not None:
        price = _number(previous.get("最新价"))
        if price is not None and price > 0:
            return price
    return None


def _trade_to_close_fraction(timestamp, report_date):
    return _remaining_trading_day_fraction(timestamp, report_date)


def _remaining_trading_day_fraction(timestamp, report_date):
    try:
        timestamp = pd.Timestamp(timestamp)
    except Exception:
        return 0.0
    day = pd.Timestamp(report_date).date()
    morning_open = pd.Timestamp(f"{day} 09:30:00")
    morning_close = pd.Timestamp(f"{day} 11:30:00")
    afternoon_open = pd.Timestamp(f"{day} 13:00:00")
    close_ts = pd.Timestamp(f"{day} 15:00:00")
    if timestamp >= close_ts:
        return 0.0
    if timestamp <= morning_open:
        return 1.0
    if timestamp < morning_close:
        minutes = (
            (morning_close - timestamp).total_seconds() / 60.0
            + 120.0
        )
    elif timestamp < afternoon_open:
        minutes = 120.0
    else:
        minutes = (close_ts - timestamp).total_seconds() / 60.0
    return min(max(minutes / 240.0, 0.0), 1.0)


def _trading_day_steps(dates, product=None):
    timestamps = pd.to_datetime(dates, errors="coerce")
    previous = timestamps.shift(1)
    steps = pd.Series(0.0, index=dates.index, dtype=float)
    valid = previous.notna() & timestamps.notna()
    calendar = market_data.load_live_trading_calendar()
    for index in steps.index[valid]:
        previous_date = previous.loc[index].normalize()
        current_date = timestamps.loc[index].normalize()
        if len(calendar) and calendar.max() >= current_date:
            count = ((calendar > previous_date) & (calendar <= current_date)).sum()
        else:
            count = np.busday_count(previous_date.date(), current_date.date())
        steps.loc[index] = float(count)
    return steps


def _option_position_rows_for_intraday(position_history, report_date, account_id):
    rows = position_history[
        position_history["日期"].astype(str).eq(str(report_date))
        & position_history["账户ID"].astype(str).eq(str(account_id))
        & ~position_history.get("方向", pd.Series(dtype=object)).astype(str).eq("hedge")
    ].copy()
    if rows.empty:
        return rows
    rows["_intraday_leg"] = rows.apply(_position_row_leg, axis=1)
    return rows


def _segmented_hedge_delta_pnl_series(product, group, spot, hedge_mark, hedge_qty):
    price = hedge_mark.where(hedge_mark.notna(), spot)
    return (hedge_qty.shift(1).fillna(0.0) * price.diff()).fillna(0.0)


def _previous_close_hedge_pnl_series(raw_daily_pnl, spot, hedge_mark, hedge_qty):
    price = hedge_mark.where(hedge_mark.notna(), spot)
    result = (hedge_qty.shift(1).fillna(0.0) * price.diff()).fillna(0.0)
    if result.empty:
        return result
    first_raw = raw_daily_pnl.iloc[0] if len(raw_daily_pnl) else np.nan
    if pd.notna(first_raw):
        result.iloc[0] = first_raw
    return result


def _previous_close_option_pnl_series(raw_daily_pnl, group, position_history, product):
    result = pd.Series(np.nan, index=group.index, dtype="float64")
    if (
        product is None
        or position_history is None
        or position_history.empty
        or group.empty
    ):
        return result
    required = {"日期", "账户ID", "方向", "合约代码", "总持仓", "最新价"}
    if not required.issubset(position_history.columns):
        return result

    multiplier = _contract_multiplier(product)
    ordered = list(group.index)
    for offset, current_index in enumerate(ordered):
        if offset == 0:
            if len(raw_daily_pnl) and pd.notna(raw_daily_pnl.loc[current_index]):
                result.loc[current_index] = raw_daily_pnl.loc[current_index]
            continue

        current_row = group.loc[current_index]
        previous_row = group.loc[ordered[offset - 1]]
        account_id = str(current_row.get("账户ID"))
        current_date = str(current_row.get("日期"))
        previous_date = str(previous_row.get("日期"))
        previous_positions = _option_position_rows_for_intraday(
            position_history,
            previous_date,
            account_id,
        )
        if previous_positions.empty:
            result.loc[current_index] = 0.0
            continue
        current_positions = _option_position_rows_for_intraday(
            position_history,
            current_date,
            account_id,
        )

        daily_pnl = 0.0
        matched_all = True
        for _, previous_position in previous_positions.iterrows():
            code = _security_code(previous_position.get("合约代码"))
            previous_price = _number(previous_position.get("最新价"))
            if code is None or previous_price is None:
                matched_all = False
                break
            current_price = _current_option_close_price_for_previous_position(
                product,
                current_date,
                code,
                current_positions,
            )
            if current_price is None:
                matched_all = False
                break
            qty = abs(_number(previous_position.get("总持仓")) or 0.0)
            direction = (
                -1.0
                if str(previous_position.get("方向") or "").lower() == "short"
                else 1.0
            )
            daily_pnl += (
                direction
                * qty
                * (float(current_price) - float(previous_price))
                * multiplier
            )

        if matched_all:
            result.loc[current_index] = float(daily_pnl)

    return result


def _current_option_close_price_for_previous_position(
    product,
    report_date,
    code,
    current_positions,
):
    if current_positions is not None and not current_positions.empty:
        for _, row in current_positions.iterrows():
            if _security_code(row.get("合约代码")) != code:
                continue
            price = _number(row.get("最新价"))
            if price is not None and price > 0:
                return price
    return _option_close_price_from_quote_snapshot(product, report_date, code)


def _option_close_price_from_quote_snapshot(product, report_date, code):
    if product is None:
        return None
    timestamp = pd.Timestamp(f"{pd.Timestamp(report_date).date()} 23:59:59")
    return _option_price_from_quote_snapshot(product, report_date, code, timestamp)


def _current_option_iv_for_previous_position(
    product,
    report_date,
    code,
    previous_position,
    current_spot,
    current_report,
    current_positions,
):
    if current_report is not None and not current_report.empty:
        current_matches = current_report.loc[current_report["_code"].eq(code)]
        if not current_matches.empty:
            iv = _number(current_matches.iloc[0].get("IV"))
            if iv is not None and iv > 0:
                return iv
    if current_positions is not None and not current_positions.empty:
        for _, row in current_positions.iterrows():
            if _security_code(row.get("合约代码")) != code:
                continue
            iv = _number(row.get("IV"))
            if iv is not None and iv > 0:
                return iv

    price = _option_close_price_from_quote_snapshot(product, report_date, code)
    spot = _number(current_spot)
    if price is None or spot is None:
        return None
    leg = _position_row_leg(previous_position)
    flag = "c" if leg == "Call" else "p" if leg == "Put" else None
    if flag is None:
        return None
    qty = abs(_number(previous_position.get("总持仓")) or 0.0)
    if qty <= 0:
        return None
    direction = -1.0 if str(previous_position.get("方向") or "").lower() == "short" else 1.0
    close_dte = _option_close_dte(
        product,
        report_date,
        {"previous": previous_position},
        previous_position,
    )
    if close_dte is None or close_dte <= 0:
        return None
    greeks = _option_greeks_for_dte(
        product,
        previous_position,
        price,
        spot,
        flag,
        direction * qty,
        close_dte,
        fallback_iv=_number(previous_position.get("IV")),
    )
    if greeks is None:
        return None
    return greeks["iv"]


def _segmented_hedge_delta_pnl(previous_qty, start_price, end_price, trade_rows):
    pnl = 0.0
    qty = float(previous_qty or 0.0)
    last_price = float(start_price)
    for row in trade_rows:
        trade_price = _number(row.get("price"))
        signed_qty = _number(row.get("signed_qty"))
        if trade_price is None or signed_qty is None:
            continue
        pnl += qty * (trade_price - last_price)
        qty += signed_qty
        last_price = trade_price
    return pnl + qty * (float(end_price) - last_price)


def _security_trade_rows_by_date(product, account_id="default"):
    rows_by_date = {}
    for row in account_store.list_fills(
        product,
        account_id=account_id,
        include_voided=False,
        order="asc",
    ):
        fill = _supported_fill(row["payload"])
        if fill is None:
            continue
        action = str(fill.get("action") or "").lower()
        if action not in {"delta_hedge", "rebalance_hedge", "close_hedge"}:
            continue
        trade_date = str(fill.get("date") or "")
        rows_by_date.setdefault(trade_date, []).extend(_hedge_fill_trade_events(fill))
    return rows_by_date


def _fill_summary_leg_greeks_from_positions(summary_history, position_history, product=None):
    if position_history is None or position_history.empty:
        return summary_history
    multiplier = _contract_multiplier(product)
    aggregates = {}
    for _, row in position_history.iterrows():
        leg = _position_row_leg(row)
        if leg is None:
            continue
        key = (str(row.get("日期")), str(row.get("账户ID")))
        target = aggregates.setdefault(
            key,
            {
                "direct": _empty_leg_aggregate(),
                "scaled": _empty_leg_aggregate(),
            },
        )
        qty = abs(_number(row.get("总持仓")) or 0.0)
        scale = qty * multiplier
        direction = -1.0 if str(row.get("方向") or "") == "short" else 1.0
        for metric in ["Delta", "Gamma", "Vega", "Theta"]:
            value = _number(row.get(metric))
            if metric == "Delta":
                single_delta = _number(row.get("单张Delta"))
                if single_delta is not None:
                    value = single_delta * scale
            if value is not None:
                target["direct"][f"{leg} {metric}"] += value
                target["scaled"][f"{leg} {metric}"] += value * scale * direction
        iv = _number(row.get("IV"))
        if iv is not None:
            target["direct"][f"{leg} IV"].append(iv)
            target["scaled"][f"{leg} IV"].append(iv)

    if not aggregates:
        return summary_history

    result = summary_history.copy()
    for index, row in result.iterrows():
        values = aggregates.get((str(row.get("日期")), str(row.get("账户ID"))))
        if values is None:
            continue
        selected = values[_select_leg_aggregate_mode(row, values)]
        replace_greeks = _should_replace_leg_greeks(row, selected) or (
            _selected_leg_iv_differs(row, selected)
        )
        for column, value in selected.items():
            if column.endswith(" IV"):
                if value:
                    value = sum(value) / len(value)
                else:
                    continue
            if column not in result.columns:
                result[column] = None
            if pd.isna(result.at[index, column]) or (
                replace_greeks
            ):
                result.at[index, column] = value
    return result


def _backfill_position_single_delta_columns(position_history, product=None):
    if position_history is None or position_history.empty:
        return position_history
    result = position_history.copy()
    for column in ["单张Delta", "Delta"]:
        if column not in result.columns:
            result[column] = None
    multiplier_default = _contract_multiplier(product)
    for index, row in result.iterrows():
        if str(row.get("方向") or "").lower() == "hedge":
            result.at[index, "单张Delta"] = None
            continue
        qty = abs(_number(row.get("总持仓")) or 0.0)
        if qty <= 0:
            continue
        multiplier = _number(row.get("合约乘数")) or multiplier_default
        scale = qty * multiplier
        if scale <= 0:
            continue
        side = str(row.get("方向") or "").lower()
        direction = -1.0 if side == "short" else 1.0
        single_delta = _number(row.get("单张Delta"))
        delta = _number(row.get("Delta"))

        if single_delta is None and delta is not None:
            if abs(delta) <= 5.0:
                single_delta = direction * delta
            else:
                single_delta = delta / scale

        if single_delta is None:
            continue
        result.at[index, "单张Delta"] = single_delta
        result.at[index, "Delta"] = single_delta * scale
    return result


def _fill_summary_hedge_marks_and_pnl(summary_history, position_history, product=None):
    result = summary_history.copy()
    if position_history is not None and not position_history.empty:
        hedge_rows = position_history[
            position_history.get("方向", pd.Series(dtype=object)).astype(str).eq("hedge")
        ]
        for _, hedge_row in hedge_rows.iterrows():
            mask = (
                result["日期"].astype(str).eq(str(hedge_row.get("日期")))
                & result["账户ID"].astype(str).eq(str(hedge_row.get("账户ID")))
            )
            if mask.any():
                result.loc[mask, "对冲最新价"] = _number(hedge_row.get("最新价"))
    legacy_columns = {
        "对冲成本",
        "对冲浮盈亏",
        "对冲已实现盈亏",
        "对冲总盈亏",
        "期权已实现盈亏",
        "期权总盈亏",
        "估算权益",
    }
    if product is None or not legacy_columns.intersection(result.columns):
        return result
    for column in legacy_columns:
        if column not in result.columns:
            result[column] = None

    for index, row in result.iterrows():
        account_id = str(row.get("账户ID"))
        report_date = str(row.get("日期"))
        hedge_qty = _number(row.get("对冲持仓")) or 0.0
        hedge_cost = _hedge_open_cost_for_report(product, account_id, report_date)
        if hedge_cost is None:
            hedge_cost = _number(row.get("对冲成本")) or 0.0
        hedge_mark = _number(row.get("对冲最新价"))
        hedge_unrealized = (
            core.hedge.calc_unrealized_pnl(hedge_qty, hedge_cost, hedge_mark)
            if hedge_mark is not None and hedge_cost > 0
            else 0.0
        )
        hedge_realized = _cumulative_hedge_realized_pnl_for_report(
            product,
            account_id,
            report_date,
        )
        option_realized = _cumulative_option_realized_pnl_for_report(
            product,
            account_id,
            report_date,
        )
        option_unrealized = _number(row.get("期权浮盈亏")) or 0.0
        if "对冲成本" in result.columns:
            result.at[index, "对冲成本"] = hedge_cost
        if "对冲浮盈亏" in result.columns:
            result.at[index, "对冲浮盈亏"] = hedge_unrealized
        if "对冲已实现盈亏" in result.columns:
            result.at[index, "对冲已实现盈亏"] = hedge_realized
        if "对冲总盈亏" in result.columns:
            result.at[index, "对冲总盈亏"] = hedge_realized + hedge_unrealized
        if "期权已实现盈亏" in result.columns:
            result.at[index, "期权已实现盈亏"] = option_realized
        if "期权总盈亏" in result.columns:
            result.at[index, "期权总盈亏"] = option_realized + option_unrealized
        if "估算权益" in result.columns:
            initial_cash = _number(row.get("初始资金")) or 0.0
            fee = _number(row.get("手续费")) or 0.0
            result.at[index, "估算权益"] = (
                initial_cash
                + option_realized
                + option_unrealized
                + hedge_realized
                + hedge_unrealized
                - fee
            )
    return result


def _empty_leg_aggregate():
    return {
        "Call Delta": 0.0,
        "Put Delta": 0.0,
        "Call Gamma": 0.0,
        "Put Gamma": 0.0,
        "Call Vega": 0.0,
        "Put Vega": 0.0,
        "Call Theta": 0.0,
        "Put Theta": 0.0,
        "Call IV": [],
        "Put IV": [],
    }


def _select_leg_aggregate_mode(summary_row, aggregates):
    expected_delta = _number(summary_row.get("期权Delta"))
    if expected_delta is not None:
        direct_delta = (
            aggregates["direct"]["Call Delta"]
            + aggregates["direct"]["Put Delta"]
        )
        scaled_delta = (
            aggregates["scaled"]["Call Delta"]
            + aggregates["scaled"]["Put Delta"]
        )
        if abs(scaled_delta - expected_delta) < abs(direct_delta - expected_delta):
            return "scaled"
        return "direct"

    direct_abs_delta = max(
        abs(aggregates["direct"]["Call Delta"]),
        abs(aggregates["direct"]["Put Delta"]),
    )
    scaled_abs_delta = max(
        abs(aggregates["scaled"]["Call Delta"]),
        abs(aggregates["scaled"]["Put Delta"]),
    )
    return "scaled" if direct_abs_delta <= 5.0 < scaled_abs_delta else "direct"


def _should_replace_leg_greeks(summary_row, selected):
    expected_delta = _number(summary_row.get("期权Delta"))
    current_call_delta = _number(summary_row.get("Call Delta"))
    current_put_delta = _number(summary_row.get("Put Delta"))
    if (
        expected_delta is None
        or current_call_delta is None
        or current_put_delta is None
    ):
        return False
    current_delta = current_call_delta + current_put_delta
    selected_delta = selected["Call Delta"] + selected["Put Delta"]
    current_error = abs(current_delta - expected_delta)
    selected_error = abs(selected_delta - expected_delta)
    tolerance = max(1.0, abs(expected_delta) * 1e-4)
    return current_error > tolerance and selected_error < current_error


def _selected_leg_iv_differs(summary_row, selected):
    for leg in ("Call", "Put"):
        values = selected.get(f"{leg} IV") or []
        if not values:
            continue
        selected_iv = sum(values) / len(values)
        current_iv = _number(summary_row.get(f"{leg} IV"))
        if current_iv is None or abs(current_iv - selected_iv) > 1e-10:
            return True
    return False


def _contract_multiplier(product):
    if product is None:
        return 1.0
    try:
        return float(load_product_config(product).vol.contract_multiplier)
    except Exception:
        return 1.0


def _position_row_leg(row):
    name = str(row.get("合约名称") or "").upper()
    if "购" in name or "CALL" in name:
        return "Call"
    if "沽" in name or "PUT" in name:
        return "Put"
    direction = str(row.get("方向") or "")
    delta = _number(row.get("Delta"))
    if delta is None:
        return None
    if abs(delta) <= 1.5:
        return "Call" if delta > 0 else "Put"
    if direction == "short":
        return "Call" if delta < 0 else "Put"
    if direction == "long":
        return "Call" if delta > 0 else "Put"
    return None


def _numeric_series(frame, column):
    if column not in frame.columns:
        return pd.Series([None] * len(frame), index=frame.index, dtype="float64")
    return pd.to_numeric(frame[column], errors="coerce")


def _actual_daily_pnl_series_from_positions(
    group,
    position_history,
    product,
    side,
):
    if product is None or position_history is None or position_history.empty:
        return pd.Series(np.nan, index=group.index, dtype="float64")
    required = {"日期", "账户ID", "方向"}
    if not required.issubset(position_history.columns):
        return pd.Series(np.nan, index=group.index, dtype="float64")

    totals = []
    for _, row in group.iterrows():
        account_id = str(row.get("账户ID"))
        report_date = str(row.get("日期"))
        positions = position_history.loc[
            position_history["日期"].astype(str).eq(report_date)
            & position_history["账户ID"].astype(str).eq(account_id)
        ]
        if side == "hedge":
            hedge_qty = _number(row.get("对冲持仓")) or 0.0
            hedge_mark = _number(row.get("对冲最新价"))
            hedge_cost = _hedge_open_cost_for_report(product, account_id, report_date)
            hedge_unrealized = (
                core.hedge.calc_unrealized_pnl(hedge_qty, hedge_cost, hedge_mark)
                if hedge_mark is not None
                and hedge_cost is not None
                and hedge_cost > 0
                else 0.0
            )
            realized = _cumulative_hedge_realized_pnl_for_report(
                product,
                account_id,
                report_date,
            )
            totals.append(hedge_unrealized + realized)
            continue

        option_positions = positions.loc[
            ~positions["方向"].astype(str).str.lower().eq("hedge")
        ]
        option_unrealized = _sum_numeric_column(option_positions, "浮动盈亏")
        option_realized = _cumulative_option_realized_pnl_for_report(
            product,
            account_id,
            report_date,
        )
        totals.append(option_unrealized + option_realized)

    total_series = pd.Series(totals, index=group.index, dtype="float64")
    daily = total_series.diff()
    if not daily.empty:
        daily.iloc[0] = total_series.iloc[0]
    return daily


def _legacy_total_daily_pnl_series(group, total_column, unrealized_column):
    total = _numeric_series(group, total_column)
    unrealized = _numeric_series(group, unrealized_column)
    daily = total.diff()
    daily = daily.where(daily.notna(), unrealized.diff())
    if not daily.empty:
        daily.iloc[0] = total.iloc[0]
    return daily


def _hedge_mark_price_type(report_date):
    report_ts = _date_or_none(report_date)
    if report_ts is None:
        return None
    today = pd.Timestamp.now().date()
    if hasattr(report_ts, "date"):
        report_ts = report_ts.date()
    return "latest" if report_ts >= today else "close"


def _refresh_current_summary_from_history(payload):
    history = payload.get("summary_history")
    if history is None or history.empty:
        return
    mask = (
        history["日期"].astype(str).eq(str(payload["date"]))
        & history["账户ID"].astype(str).eq(str(payload["account_id"]))
    )
    if mask.any():
        payload["summary"] = history.loc[mask].iloc[-1].to_dict()
        _apply_current_pnl_decomposition(payload)


def _backfill_summary_financial_columns(product, account_id, summary_history):
    if summary_history is None or summary_history.empty:
        return summary_history

    result = summary_history.copy()
    reset_at = account_store.load_account(product, account_id=account_id).reset_at

    account_mask = result["账户ID"].astype(str).eq(str(account_id))
    for idx, row in result.loc[account_mask].iterrows():
        report_date = row.get("日期")
        if pd.isna(report_date):
            continue

        trade_rows = _trade_rows_from_export(
            product,
            str(report_date),
            not_before=reset_at,
        )
        trade_rows.extend(
            _etf_trade_rows_from_export(
                product,
                str(report_date),
                not_before=reset_at,
            )
        )

        if _number(row.get("当日手续费")) is None:
            result.at[idx, "当日手续费"] = _configured_daily_report_fee(
                product,
                account_id,
                str(report_date),
                trade_rows,
            )

    return result.reindex(columns=SUMMARY_COLUMNS)


def _configured_daily_report_fee(product, account_id, report_date, trade_rows):
    fee = _sum_row_values(trade_rows, "手续费")
    has_option_trade_rows = any(row.get("类型") != "ETF对冲" for row in trade_rows)
    has_etf_trade_rows = any(row.get("类型") == "ETF对冲" for row in trade_rows)

    fills = account_store.list_fills(
        product,
        account_id=account_id,
        include_voided=False,
    )
    for row in fills:
        fill = _supported_fill(row["payload"])
        if fill is None:
            continue
        if str(fill.get("date")) != str(report_date):
            continue
        action = str(fill.get("action") or "").lower()
        if action in {
            "open_long_straddle",
            "open_short_straddle",
            "close_long_straddle",
            "close_short_straddle",
            "roll_long_straddle",
            "roll_short_straddle",
        }:
            if not has_option_trade_rows and not _is_position_only_holding_import(fill):
                fee += _configured_option_fill_fee(product, fill)
        elif action in {"delta_hedge", "rebalance_hedge", "close_hedge"}:
            if not has_etf_trade_rows:
                fee += _configured_hedge_fill_fee(product, fill)
    return fee


def _configured_cumulative_report_fee(product, account_id, report_date):
    cutoff = _date_or_none(report_date)
    if cutoff is None:
        return 0.0

    live_account = account_store.load_account(product, account_id=account_id)
    option_rows = _all_trade_rows_from_exports(product, not_before=live_account.reset_at)
    etf_rows = _all_etf_trade_rows_from_exports(product, not_before=live_account.reset_at)
    fee = 0.0
    option_trade_dates = set()
    etf_trade_dates = set()
    for row in option_rows:
        row_date = _date_or_none(row.get("日期"))
        if row_date is None or row_date > cutoff:
            continue
        option_trade_dates.add(str(row_date.date()))
        fee += float(row.get("手续费") or 0.0)
    for row in etf_rows:
        row_date = _date_or_none(row.get("日期"))
        if row_date is None or row_date > cutoff:
            continue
        etf_trade_dates.add(str(row_date.date()))
        fee += float(row.get("手续费") or 0.0)
    fills = account_store.list_fills(
        product,
        account_id=account_id,
        include_voided=False,
    )
    for row in fills:
        fill = _supported_fill(row["payload"])
        if fill is None:
            continue
        fill_date = _date_or_none(fill.get("date"))
        if fill_date is None or fill_date > cutoff:
            continue
        date_key = str(fill_date.date())
        action = str(fill.get("action") or "").lower()
        if action in {
            "open_long_straddle",
            "open_short_straddle",
            "close_long_straddle",
            "close_short_straddle",
            "roll_long_straddle",
            "roll_short_straddle",
        }:
            if date_key not in option_trade_dates and not _is_position_only_holding_import(fill):
                fee += _configured_option_fill_fee(product, fill)
        elif action in {"delta_hedge", "rebalance_hedge", "close_hedge"}:
            if date_key not in etf_trade_dates:
                fee += _configured_hedge_fill_fee(product, fill)
    return fee


def _cumulative_option_realized_pnl_for_report(product, account_id, report_date):
    cutoff = _date_or_none(report_date)
    if cutoff is None:
        return 0.0
    realized = 0.0
    positions = {}
    fills = account_store.list_fills(
        product,
        account_id=account_id,
        include_voided=False,
    )
    for row in fills:
        fill = _supported_fill(row["payload"])
        if fill is None:
            continue
        fill_date = _date_or_none(fill.get("date"))
        if fill_date is None or fill_date > cutoff:
            continue
        action = str(fill.get("action") or "").lower()
        side = str(fill.get("side") or _side_from_option_action(action) or "")
        if action in {"open_long_straddle", "open_short_straddle", "open_straddle"}:
            positions[side or "long"] = fill
            continue
        if action in {"roll_long_straddle", "roll_short_straddle", "roll_straddle"}:
            explicit = _number(fill.get("realized_pnl"))
            if explicit is not None:
                realized += explicit
            positions[side or "short"] = fill
            continue
        if action in {"close_long_straddle", "close_short_straddle", "close_straddle"}:
            explicit = _number(fill.get("realized_pnl"))
            if explicit is not None:
                realized += explicit
            else:
                position = positions.get(side)
                realized += _straddle_close_realized_pnl(position, fill)
            positions[side] = None
            continue
        if action == "rebalance_straddle_legs":
            explicit = _number(fill.get("realized_pnl"))
            if explicit is not None:
                realized += explicit
            else:
                realized += _straddle_rebalance_realized_pnl(
                    positions.get(side),
                    fill,
                )
            positions[side] = fill
            continue
    return realized


def _side_from_option_action(action):
    if "short" in str(action):
        return "short"
    if "long" in str(action):
        return "long"
    return None


def _straddle_close_realized_pnl(position, close_fill):
    if not position:
        return 0.0
    side = str(close_fill.get("side") or position.get("side") or "")
    sign = -1.0 if side == "short" else 1.0
    multiplier = float(
        _number(
            close_fill.get(
                "contract_multiplier",
                position.get("contract_multiplier"),
            )
        )
        or 1.0
    )
    total = 0.0
    close_prices = _close_leg_prices(close_fill)
    for leg in ("call", "put"):
        qty = min(
            float(_number(position.get(f"{leg}_qty")) or 0.0),
            float(_number(close_fill.get(f"{leg}_qty")) or 0.0),
        )
        if qty <= 0:
            continue
        entry_price = _number(position.get(f"entry_{leg}_price"))
        close_price = close_prices.get(leg)
        if entry_price is None or close_price is None:
            continue
        total += qty * (float(close_price) - float(entry_price)) * sign * multiplier
    return total


def _straddle_rebalance_realized_pnl(position, rebalance_fill):
    if not position:
        return 0.0
    side = str(rebalance_fill.get("side") or position.get("side") or "")
    sign = -1.0 if side == "short" else 1.0
    multiplier = float(
        _number(
            rebalance_fill.get(
                "contract_multiplier",
                position.get("contract_multiplier"),
            )
        )
        or 1.0
    )
    total = 0.0
    for adjustment in rebalance_fill.get("leg_adjustments") or []:
        leg = str(adjustment.get("leg") or "").lower()
        if leg not in {"call", "put"}:
            continue
        qty_change = _number(adjustment.get("qty_change"))
        qty = (
            abs(qty_change)
            if qty_change is not None
            else _number(adjustment.get("qty"))
        )
        if qty is None or qty <= 0:
            continue
        previous_qty = _number(position.get(f"{leg}_qty")) or 0.0
        current_qty = _number(rebalance_fill.get(f"{leg}_qty")) or 0.0
        if qty_change is None:
            qty = min(float(qty), max(0.0, previous_qty - current_qty))
        elif qty_change >= 0:
            continue
        entry_price = _number(position.get(f"entry_{leg}_price"))
        close_price = _number(adjustment.get("price"))
        if entry_price is None or close_price is None:
            continue
        total += (
            float(qty)
            * (float(close_price) - float(entry_price))
            * sign
            * multiplier
        )
    return total


def _close_leg_prices(close_fill):
    result = {
        "call": _number(close_fill.get("call_price")),
        "put": _number(close_fill.get("put_price")),
    }
    for leg_close in close_fill.get("leg_closes") or []:
        leg = str(leg_close.get("leg") or "").lower()
        if leg in result and result[leg] is None:
            result[leg] = _number(leg_close.get("price"))
    return result


def _refresh_current_summary_greeks_from_position_rows(payload):
    rows = _frame(payload.get("position_rows", []), POSITION_COLUMNS)
    if rows.empty or "方向" not in rows.columns:
        return
    option_rows = rows.loc[~rows["方向"].astype(str).str.lower().eq("hedge")].copy()
    if option_rows.empty:
        return

    summary = payload.get("summary")
    if not isinstance(summary, dict):
        return

    leg_values = {
        "call": {"iv": [], "delta": 0.0, "gamma": 0.0, "vega": 0.0, "theta": 0.0},
        "put": {"iv": [], "delta": 0.0, "gamma": 0.0, "vega": 0.0, "theta": 0.0},
    }
    totals = {"delta": 0.0, "gamma": 0.0, "vega": 0.0, "theta": 0.0}
    position_ivs = []
    for _, row in option_rows.iterrows():
        leg = _position_row_leg(row)
        leg = None if leg is None else str(leg).lower()
        if leg not in leg_values:
            continue
        iv = _number(row.get("IV"))
        if iv is not None:
            leg_values[leg]["iv"].append(iv)
            position_ivs.append(iv)
        for target, column in [
            ("delta", "Delta"),
            ("gamma", "Gamma"),
            ("vega", "Vega"),
            ("theta", "Theta"),
        ]:
            value = _number(row.get(column))
            if value is None:
                continue
            leg_values[leg][target] += value
            totals[target] += value

    hedge_qty = _number(summary.get("对冲持仓")) or 0.0
    summary["期权Delta"] = totals["delta"]
    summary["账户Delta"] = totals["delta"] + hedge_qty
    summary["账户Gamma"] = totals["gamma"]
    summary["账户Vega"] = totals["vega"]
    summary["账户Theta"] = totals["theta"]
    if position_ivs:
        summary["持仓IV"] = sum(position_ivs) / len(position_ivs)
    for leg, prefix in [("call", "Call"), ("put", "Put")]:
        values = leg_values[leg]
        if values["iv"]:
            summary[f"{prefix} IV"] = sum(values["iv"]) / len(values["iv"])
        summary[f"{prefix} Delta"] = values["delta"]
        summary[f"{prefix} Gamma"] = values["gamma"]
        summary[f"{prefix} Vega"] = values["vega"]
        summary[f"{prefix} Theta"] = values["theta"]


def _repair_zero_iv_position_rows_with_intraday_minutes(position_rows, product):
    """Use same-minute ETF/option prices to repair current-day zero IV rows."""
    if position_rows is None or product is None:
        return position_rows
    as_records = not isinstance(position_rows, pd.DataFrame)
    result = _frame(position_rows, POSITION_COLUMNS) if as_records else position_rows.copy()
    if result.empty:
        return [] if as_records else result
    required = {"日期", "方向", "合约代码", "总持仓", "最新价", "行权价", "到期日", "IV"}
    if not required.issubset(result.columns):
        return position_rows

    spec = market_data.SSE_ETF_OPTION_SPECS.get(product)
    if spec is None:
        return position_rows
    calendar = market_data.load_live_trading_calendar()
    load_product_config(product)
    quote_cache = {}
    option_mask = ~result["方向"].astype(str).str.lower().eq("hedge")
    zero_iv_mask = result["IV"].apply(_number).fillna(0.0).le(0.0)
    for index, row in result.loc[option_mask & zero_iv_mask].iterrows():
        code = _security_code(row.get("合约代码"))
        report_date = _date_or_none(row.get("日期"))
        leg = _position_row_leg(row)
        leg = None if leg is None else str(leg).lower()
        strike = _number(row.get("行权价"))
        maturity = row.get("到期日")
        if code is None or report_date is None or leg is None or strike is None or maturity is None:
            continue

        date_text = str(pd.Timestamp(report_date).date())
        cache_key = (date_text, code)
        if cache_key not in quote_cache:
            quote_cache[cache_key] = _latest_positive_intraday_option_quote(
                product,
                spec.etf_symbol,
                date_text,
                code,
                leg,
                strike,
                maturity,
                calendar,
            )
        quote = quote_cache[cache_key]
        if quote is None:
            continue

        qty = abs(_number(row.get("总持仓")) or 0.0)
        if qty <= 0:
            continue
        multiplier = _contract_multiplier(product)
        direction = -1.0 if str(row.get("方向") or "").lower() == "short" else 1.0
        scale = direction * qty * multiplier
        result.at[index, "剩余天数"] = quote.get("dte")
        result.at[index, "IV"] = quote.get("iv")
        delta = _number(quote.get("delta"))
        result.at[index, "单张Delta"] = None if delta is None else direction * delta
        for column, source in [
            ("Delta", "delta"),
            ("Gamma", "gamma"),
            ("Vega", "vega"),
            ("Theta", "theta"),
        ]:
            value = _number(quote.get(source))
            result.at[index, column] = None if value is None else value * scale

    if as_records:
        return result.to_dict("records")
    return result


def _latest_positive_intraday_option_quote(
    product,
    etf_symbol,
    report_date,
    option_code,
    leg,
    strike,
    maturity,
    trading_calendar,
):
    intraday_dir = (
        storage.PROJECT_ROOT
        / "data"
        / "live"
        / product
        / "intraday"
        / pd.Timestamp(report_date).strftime("%Y%m%d")
    )
    option_path = intraday_dir / f"option_{option_code}_1m.csv"
    etf_path = intraday_dir / f"etf_{etf_symbol}_1m.csv"
    option, etf = _load_intraday_minute_frames_for_iv_repair(
        option_path,
        etf_path,
        etf_symbol,
        report_date,
        option_code,
    )
    if option is None or etf is None:
        return None
    if not {"timestamp", "price"}.issubset(option.columns):
        option = None
    if not {"timestamp", "close"}.issubset(etf.columns):
        etf = None
    quote = _positive_intraday_option_quote_from_minute_frames(
        product,
        report_date,
        option_code,
        leg,
        strike,
        maturity,
        trading_calendar,
        option,
        etf,
    )
    if quote is not None:
        return quote
    if option_path.exists() and etf_path.exists():
        option, etf = _fetch_akshare_intraday_minute_frames_for_iv_repair(
            etf_symbol,
            report_date,
            option_code,
        )
        return _positive_intraday_option_quote_from_minute_frames(
            product,
            report_date,
            option_code,
            leg,
            strike,
            maturity,
            trading_calendar,
            option,
            etf,
        )
    return None


def _positive_intraday_option_quote_from_minute_frames(
    product,
    report_date,
    option_code,
    leg,
    strike,
    maturity,
    trading_calendar,
    option,
    etf,
):
    if option is None or etf is None:
        return None
    if not {"timestamp", "price"}.issubset(option.columns):
        return None
    if not {"timestamp", "close"}.issubset(etf.columns):
        return None

    target_date = pd.Timestamp(report_date).date()
    option = option.copy()
    etf = etf.copy()
    option["timestamp"] = pd.to_datetime(option["timestamp"], errors="coerce")
    etf["timestamp"] = pd.to_datetime(etf["timestamp"], errors="coerce")
    option["price"] = pd.to_numeric(option["price"], errors="coerce")
    etf["close"] = pd.to_numeric(etf["close"], errors="coerce")
    option = option.dropna(subset=["timestamp", "price"])
    etf = etf.dropna(subset=["timestamp", "close"])
    option = option.loc[option["timestamp"].dt.date.eq(target_date)]
    etf = etf.loc[etf["timestamp"].dt.date.eq(target_date)]
    if option.empty or etf.empty:
        return None

    option["minute"] = option["timestamp"].dt.floor("min")
    etf["minute"] = etf["timestamp"].dt.floor("min")
    option = option.sort_values("timestamp").drop_duplicates("minute", keep="last")
    etf = etf.sort_values("timestamp").drop_duplicates("minute", keep="last")
    merged = option[["minute", "price"]].merge(
        etf[["minute", "close"]],
        on="minute",
        how="inner",
    )
    close_time = pd.Timestamp(f"{target_date} 15:00:00")
    merged = merged.loc[merged["minute"].le(close_time)]
    if merged.empty:
        return None

    flag = "c" if leg == "call" else "p"
    for _, minute_row in merged.sort_values("minute", ascending=False).iterrows():
        price = _number(minute_row.get("price"))
        spot = _number(minute_row.get("close"))
        if price is None or spot is None or price <= 0 or spot <= 0:
            continue
        chain = pd.DataFrame(
            [
                {
                    "date": report_date,
                    "order_book_id": option_code,
                    "maturity_date": maturity,
                    "strike_price": strike,
                    "option_type": flag,
                    "bid": price,
                    "ask": price,
                    "volume": 1,
                    "contract_multiplier": _contract_multiplier(product),
                    "underlying_close": spot,
                }
            ]
        )
        try:
            chain = core.vol_engine.add_iv_for_day(
                chain,
                spot,
                trading_calendar=trading_calendar,
            )
            iv = _number(chain.iloc[0].get("iv"))
            if iv is None or iv <= 0:
                continue
            chain = core.vol_engine.add_greeks_for_day(chain, spot)
        except Exception:
            continue
        quote = chain.iloc[0].to_dict()
        quote["intraday_timestamp"] = minute_row.get("minute")
        quote["intraday_option_price"] = price
        quote["intraday_spot"] = spot
        return quote
    return None


def _load_intraday_minute_frames_for_iv_repair(
    option_path,
    etf_path,
    etf_symbol,
    report_date,
    option_code,
):
    if option_path.exists() and etf_path.exists():
        try:
            return (
                pd.read_csv(option_path, encoding="utf-8-sig"),
                pd.read_csv(etf_path, encoding="utf-8-sig"),
            )
        except (OSError, ValueError):
            return None, None
    return _fetch_akshare_intraday_minute_frames_for_iv_repair(
        etf_symbol,
        report_date,
        option_code,
    )


def _fetch_akshare_intraday_minute_frames_for_iv_repair(
    etf_symbol,
    report_date,
    option_code,
):
    try:
        import akshare as ak
    except ImportError:
        return None, None

    target_date = pd.Timestamp(report_date).date()
    option = _fetch_akshare_option_minute_for_iv_repair(
        ak,
        option_code,
        target_date,
    )
    etf = _fetch_akshare_etf_minute_for_iv_repair(
        ak,
        etf_symbol,
        target_date,
    )
    if option is None or etf is None:
        return None, None
    return option, etf


def _fetch_akshare_option_minute_for_iv_repair(ak, option_code, target_date):
    frames = []
    try:
        frames.append(ak.option_finance_minute_sina(symbol=str(option_code)))
    except Exception:
        pass
    try:
        frames.append(ak.option_sse_minute_sina(symbol=str(option_code)))
    except Exception:
        pass

    for raw in frames:
        parsed = _parse_option_minute_frame_for_iv_repair(raw)
        if parsed is None:
            continue
        parsed = parsed.loc[parsed["timestamp"].dt.date.eq(target_date)]
        if not parsed.empty:
            return parsed
    return None


def _parse_option_minute_frame_for_iv_repair(raw):
    if raw is None or raw.empty:
        return None
    frame = raw.copy()
    columns = set(frame.columns)
    if {"timestamp", "price"}.issubset(columns):
        frame["timestamp"] = pd.to_datetime(frame["timestamp"], errors="coerce")
        frame["price"] = pd.to_numeric(frame["price"], errors="coerce")
        return frame[["timestamp", "price"]].dropna(subset=["timestamp", "price"])
    if {"date", "time"}.issubset(columns):
        frame["timestamp"] = pd.to_datetime(
            frame["date"].astype(str) + " " + frame["time"].astype(str),
            errors="coerce",
        )
        price_column = "price"
    elif {"日期", "时间"}.issubset(columns):
        frame["timestamp"] = pd.to_datetime(
            frame["日期"].astype(str) + " " + frame["时间"].astype(str),
            errors="coerce",
        )
        price_column = "价格"
    else:
        return None
    if price_column not in frame.columns:
        return None
    frame["price"] = pd.to_numeric(frame[price_column], errors="coerce")
    return frame[["timestamp", "price"]].dropna(subset=["timestamp", "price"])


def _fetch_akshare_etf_minute_for_iv_repair(ak, etf_symbol, target_date):
    frames = []
    try:
        frames.append(ak.stock_zh_a_minute(symbol=f"sh{etf_symbol}", period="1", adjust=""))
    except Exception:
        pass
    try:
        frames.append(ak.fund_etf_hist_min_em(symbol=etf_symbol, period="1", adjust=""))
    except Exception:
        pass

    for raw in frames:
        parsed = _parse_etf_minute_frame_for_iv_repair(raw)
        if parsed is None:
            continue
        parsed = parsed.loc[parsed["timestamp"].dt.date.eq(target_date)]
        if not parsed.empty:
            return parsed
    return None


def _parse_etf_minute_frame_for_iv_repair(raw):
    if raw is None or raw.empty:
        return None
    frame = raw.copy()
    columns = set(frame.columns)
    timestamp_column = None
    for candidate in ["timestamp", "day", "时间", "日期时间"]:
        if candidate in columns:
            timestamp_column = candidate
            break
    if timestamp_column is None:
        timestamp_column = frame.columns[0]

    close_column = None
    for candidate in ["close", "收盘", "最新价", "价格"]:
        if candidate in columns:
            close_column = candidate
            break
    if close_column is None:
        return None

    frame["timestamp"] = pd.to_datetime(frame[timestamp_column], errors="coerce")
    frame["close"] = pd.to_numeric(frame[close_column], errors="coerce")
    return frame[["timestamp", "close"]].dropna(subset=["timestamp", "close"])


def _revalue_stale_position_greeks(position_history, product):
    """Repair historical IV/Greeks when a newer calendar changes stored DTE."""
    if position_history is None or position_history.empty or product is None:
        return position_history
    required = {"日期", "方向", "合约代码", "总持仓", "到期日", "剩余天数"}
    if not required.issubset(position_history.columns):
        return position_history

    calendar = market_data.load_live_trading_calendar()
    if len(calendar) == 0:
        return position_history
    result = position_history.copy()
    option_mask = ~result["方向"].astype(str).str.lower().eq("hedge")
    stale_dates = set()
    expected_dte = {}
    for index, row in result.loc[option_mask].iterrows():
        try:
            date = pd.Timestamp(row["日期"]).normalize()
            maturity = pd.Timestamp(row["到期日"]).normalize()
        except (TypeError, ValueError):
            continue
        dte = core.vol_engine._count_trading_dte(
            date,
            maturity,
            trading_calendar=calendar,
        )
        expected_dte[index] = dte
        stored_dte = _number(row.get("剩余天数"))
        if stored_dte is None or abs(stored_dte - dte) > 1e-9:
            stale_dates.add(str(date.date()))

    if not stale_dates:
        return result

    load_product_config(product)
    for date_text in sorted(stale_dates):
        try:
            snapshot = market_data.load_latest_quote_snapshot(product, date_text)
            etf = pd.read_parquet(snapshot["etf_snapshot"])
            chain = pd.read_parquet(snapshot["option_snapshot"]).copy()
            spot = float(pd.to_numeric(etf["close"], errors="coerce").dropna().iloc[-1])
            if "date" not in chain.columns:
                chain.insert(0, "date", pd.Timestamp(date_text))
            else:
                chain["date"] = pd.Timestamp(date_text)
            chain["order_book_id"] = chain["order_book_id"].astype(str)
            chain = core.vol_engine.add_iv_for_day(
                chain,
                spot,
                trading_calendar=calendar,
            )
            chain = core.vol_engine.add_greeks_for_day(chain, spot)
        except (FileNotFoundError, OSError, KeyError, ValueError, IndexError):
            continue

        date_mask = option_mask & result["日期"].astype(str).eq(date_text)
        for index, position_row in result.loc[date_mask].iterrows():
            code = str(position_row.get("合约代码"))
            matches = chain.loc[chain["order_book_id"].eq(code)]
            if matches.empty:
                continue
            quote = matches.iloc[-1]
            qty = abs(_number(position_row.get("总持仓")) or 0.0)
            multiplier = _number(quote.get("contract_multiplier")) or _contract_multiplier(
                product
            )
            direction = (
                -1.0
                if str(position_row.get("方向") or "").lower() == "short"
                else 1.0
            )
            scale = direction * qty * multiplier
            result.at[index, "剩余天数"] = expected_dte.get(index, quote.get("dte"))
            result.at[index, "IV"] = quote.get("iv")
            delta = _number(quote.get("delta"))
            result.at[index, "单张Delta"] = (
                None if delta is None else direction * delta
            )
            for column, source in [
                ("Delta", "delta"),
                ("Gamma", "gamma"),
                ("Vega", "vega"),
                ("Theta", "theta"),
            ]:
                value = _number(quote.get(source))
                result.at[index, column] = None if value is None else value * scale
    return result


def _cumulative_hedge_realized_pnl_for_report(product, account_id, report_date):
    cutoff = _date_or_none(report_date)
    if cutoff is None:
        return 0.0

    qty = 0.0
    avg_cost = 0.0
    realized = 0.0
    fills = account_store.list_fills(
        product,
        account_id=account_id,
        include_voided=False,
    )
    for row in fills:
        fill = _supported_fill(row["payload"])
        if fill is None:
            continue
        if fill.get("action") not in {"delta_hedge", "rebalance_hedge", "close_hedge"}:
            continue
        fill_date = _date_or_none(fill.get("date"))
        if fill_date is None or fill_date > cutoff:
            continue

        trades = _hedge_fill_trade_events(fill)
        for trade in trades:
            signed_qty = _number(trade.get("signed_qty")) or 0.0
            price = _number(trade.get("price"))
            if abs(signed_qty) <= 1e-9 or price is None:
                continue
            if signed_qty > 0:
                new_qty = qty + signed_qty
                avg_cost = (
                    ((qty * avg_cost) + (signed_qty * price)) / new_qty
                    if abs(new_qty) > 1e-9
                    else 0.0
                )
                qty = new_qty
            else:
                sell_qty = min(abs(signed_qty), max(qty, 0.0))
                realized += sell_qty * (price - avg_cost)
                qty -= sell_qty
                if abs(qty) <= 1e-9:
                    qty = 0.0
                    avg_cost = 0.0

        fill_qty = _number(fill.get("qty", fill.get("new_etf_qty")))
        fill_cost = _number(fill.get("entry_price"))
        if fill_qty is not None:
            qty = fill_qty
        if fill_cost is not None and abs(qty) > 1e-9:
            avg_cost = fill_cost

    return realized


def _hedge_open_cost_for_report(product, account_id, report_date):
    cutoff = _date_or_none(report_date)
    if cutoff is None:
        return None

    qty = 0.0
    avg_cost = 0.0
    fills = account_store.list_fills(
        product,
        account_id=account_id,
        include_voided=False,
    )
    for row in fills:
        fill = _supported_fill(row["payload"])
        if fill is None:
            continue
        if fill.get("action") not in {"delta_hedge", "rebalance_hedge", "close_hedge"}:
            continue
        fill_date = _date_or_none(fill.get("date"))
        if fill_date is None or fill_date > cutoff:
            continue

        trades = _hedge_fill_trade_events(fill)
        if not trades:
            fill_qty = _number(fill.get("qty", fill.get("new_etf_qty")))
            fill_cost = _number(fill.get("entry_price"))
            if fill_qty is not None:
                qty = fill_qty
            if fill_cost is not None and qty > 0:
                avg_cost = fill_cost
            if qty <= 1e-9:
                qty = 0.0
                avg_cost = 0.0
            continue

        for trade in trades:
            signed_qty = _number(trade.get("signed_qty")) or 0.0
            price = _number(trade.get("price"))
            if abs(signed_qty) <= 1e-9 or price is None:
                continue
            if signed_qty > 0:
                new_qty = qty + signed_qty
                avg_cost = (
                    ((qty * avg_cost) + (signed_qty * price)) / new_qty
                    if new_qty > 1e-9
                    else 0.0
                )
                qty = new_qty
            else:
                qty -= min(abs(signed_qty), max(qty, 0.0))
                if qty <= 1e-9:
                    qty = 0.0
                    avg_cost = 0.0

        target_qty = _number(fill.get("qty", fill.get("new_etf_qty")))
        if target_qty is not None and abs(target_qty - qty) > 1e-6:
            qty = target_qty
            avg_cost = _number(fill.get("entry_price")) or avg_cost

    return avg_cost if qty > 1e-9 and avg_cost > 0 else None


def _hedge_fill_trade_events(fill):
    trades = []
    for trade in fill.get("security_trades") or []:
        trades.append(
            {
                "signed_qty": trade.get("signed_qty"),
                "price": trade.get("price"),
            }
        )
    if trades:
        return trades

    signed_qty = _number(fill.get("trade_etf_qty"))
    price = _number(fill.get("price"))
    if signed_qty is None or price is None or abs(signed_qty) <= 1e-9:
        return []
    return [{"signed_qty": signed_qty, "price": price}]


def _configured_option_trade_fee(product, qty):
    qty = _number(qty) or 0.0
    config = load_product_config(product)
    return abs(qty) * float(config.backtest.option_fee_per_contract)


def _configured_option_fill_fee(product, fill):
    config = load_product_config(product)
    call_qty = float(fill.get("call_qty", 0.0) or 0.0)
    put_qty = float(fill.get("put_qty", 0.0) or 0.0)
    return (abs(call_qty) + abs(put_qty)) * float(config.backtest.option_fee_per_contract)


def _configured_etf_trade_fee(product, price, qty):
    config = load_product_config(product)
    return abs(float(price or 0.0) * float(qty or 0.0)) * float(config.backtest.etf_fee_rate)


def _configured_hedge_fill_fee(product, fill):
    config = load_product_config(product)
    trades = fill.get("security_trades") or []
    if trades:
        notional = sum(
            abs(float(trade.get("price", 0.0) or 0.0) * float(trade.get("qty", 0.0) or 0.0))
            for trade in trades
        )
    else:
        qty = float(fill.get("trade_etf_qty", fill.get("qty", 0.0)) or 0.0)
        price = float(fill.get("price", fill.get("entry_price", 0.0)) or 0.0)
        notional = abs(qty * price)
    return notional * float(config.backtest.etf_fee_rate)


def _is_position_only_holding_import(fill):
    if fill.get("import_source") != "broker_holding_snapshot":
        return False
    source_file = fill.get("source_file")
    if not source_file:
        return False
    path = Path(source_file)
    if not path.exists():
        path = PROJECT_ROOT / source_file
    if not path.exists():
        return False
    try:
        df = _read_export_csv(path)
    except Exception:
        return False
    if "今开仓" not in df.columns:
        return False
    today_open = pd.to_numeric(df["今开仓"], errors="coerce").fillna(0.0).sum()
    return today_open <= 0


def _trade_rows_from_file(path, product):
    df = _read_export_csv(path)
    rows = []
    for _, item in df.iterrows():
        marker = PRODUCT_CONTRACT_NAME_MARKERS.get(product)
        if marker is not None and marker not in str(item.get("合约名称") or ""):
            continue
        row = {column: item.get(column) for column in TRADE_COLUMNS}
        row["成交时间(日)"] = item.get("成交时间(日)")
        row["策略名称"] = item.get("策略名称")
        row["买卖"] = _clean_text(row.get("买卖"))
        for column in ["报单价格", "成交价格", "成交数量", "手续费", "平仓盈亏"]:
            row[column] = _number(row.get(column))
        row["手续费"] = _configured_option_trade_fee(product, row.get("成交数量"))
        row["日期"] = _date8_to_iso(row.get("日期"))
        rows.append(row)
    return rows


def _update_history_csv(path, new_rows, columns, key_columns):
    path = Path(path)
    if path.exists():
        history = pd.read_csv(path, encoding="utf-8-sig")
    else:
        history = pd.DataFrame(columns=columns)
    incoming = _frame(new_rows, columns)
    if not incoming.empty:
        for column in columns:
            if column not in history.columns:
                history[column] = None
        mask = pd.Series(False, index=history.index)
        for _, row in incoming.iterrows():
            row_mask = pd.Series(True, index=history.index)
            for key in key_columns:
                row_mask &= history[key].astype(str) == str(row[key])
            mask |= row_mask
        history = history.loc[~mask]
        if history.empty:
            history = incoming
        else:
            history = _concat_rows([history, incoming], columns=columns)
    if "日期" in history.columns:
        history = history.sort_values([col for col in ["日期", "账户ID", "合约代码"] if col in history.columns])
    history = history.reindex(columns=columns)
    history.to_csv(path, index=False, encoding="utf-8-sig")
    return history


def _frame(rows, columns):
    return pd.DataFrame(rows, columns=columns)


def _concat_rows(frames, columns):
    rows = []
    for frame in frames:
        if frame is None or frame.empty:
            continue
        rows.extend(frame.reindex(columns=columns).to_dict("records"))
    return pd.DataFrame.from_records(rows, columns=columns)


def _option_mark_from_metadata(metadata):
    if not metadata:
        return None
    for key in ("close", "last", "latest", "price", "mid"):
        value = _number(metadata.get(key))
        if value is not None and value > 0:
            return value
    return None


def _option_mark_from_chain_row(row):
    for key in ("close", "last", "latest", "price", "mid"):
        value = _number(row.get(key))
        if value is not None and value > 0:
            return value
    bid = _number(row.get("bid"))
    ask = _number(row.get("ask"))
    if bid is not None and ask is not None and bid > 0 and ask > 0:
        return (bid + ask) / 2.0
    return bid if bid is not None and bid > 0 else ask


def _sum_row_values(rows, column):
    total = 0.0
    for row in rows:
        value = row.get(column)
        if value is None or pd.isna(value):
            continue
        total += float(value)
    return total


def _chain_metadata(chain_df):
    metadata = {}
    for _, row in chain_df.iterrows():
        code = _security_code(row.get("order_book_id"))
        mark_price = _option_mark_from_chain_row(row)
        metadata[code] = {
            "contract_symbol": row.get("contract_symbol"),
            "close": mark_price,
            "mid": mark_price,
            "strike_price": row.get("strike_price"),
            "maturity_date": str(pd.Timestamp(row.get("maturity_date")).date()),
            "dte": row.get("dte"),
            "iv": row.get("iv"),
            "delta": row.get("delta"),
            "gamma": row.get("gamma"),
            "vega": row.get("vega"),
            "theta": row.get("theta"),
            "contract_multiplier": row.get("contract_multiplier"),
        }
    return metadata


def _latest_export_file(prefix, report_date=None, not_before=None):
    files = sorted(_live_hold_dir().glob(f"{prefix}*.csv"), key=lambda path: path.stat().st_mtime)
    if report_date is not None:
        matching = [
            path
            for path in files
            if _filename_date(path) == report_date
            and _export_file_is_not_before(path, not_before)
        ]
        return matching[-1] if matching else None
    files = [path for path in files if _export_file_is_not_before(path, not_before)]
    return files[-1] if files else None


def _export_file_is_not_before(path, not_before):
    if not_before is None:
        return True
    match = re.search(
        r"(20\d{2})_(\d{2})_(\d{2})-(\d{2})_(\d{2})_(\d{2})",
        Path(path).name,
    )
    if match is None:
        return False
    timestamp = pd.Timestamp(
        "-".join(match.groups()[:3]) + " " + ":".join(match.groups()[3:])
    ).tz_localize("Asia/Hong_Kong")
    cutoff = pd.Timestamp(not_before)
    if cutoff.tzinfo is None:
        cutoff = cutoff.tz_localize("UTC")
    return timestamp.tz_convert("UTC") >= cutoff.tz_convert("UTC")


def _live_hold_dir():
    return PROJECT_ROOT / "live_hold"


def _read_export_csv(path):
    for encoding in ["utf-8-sig", "gbk", "gb18030"]:
        try:
            return pd.read_csv(path, encoding=encoding)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path)


def _filename_date(path):
    match = re.search(r"(20\d{2})_(\d{2})_(\d{2})", Path(path).name)
    if not match:
        return None
    return "-".join(match.groups())


def _date8_to_iso(value):
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if re.fullmatch(r"20\d{6}", text):
        return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    try:
        return str(pd.Timestamp(text).date())
    except Exception:
        return text


def _date_or_none(value):
    if value is None or pd.isna(value):
        return None
    try:
        return pd.Timestamp(value).normalize()
    except Exception:
        return None


def _security_code_from_underlying(value):
    if value is None:
        return None
    return _security_code(str(value).split(".", 1)[0])


def _security_code(value):
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    if text.isdigit():
        return text.zfill(6)
    return text


def _clean_text(value):
    if value is None or pd.isna(value):
        return None
    return str(value).strip()


def _supported_fill(payload):
    try:
        return account_store.normalize_fill(payload)
    except ValueError:
        return None


def _number(value):
    if value is None or pd.isna(value):
        return None
    text = str(value).strip().replace(",", "").replace("%", "")
    if text in {"", "--", "待设置", "全部"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _plain_table(rows, columns):
    if not rows:
        return ["(none)"]
    widths = {
        column: min(
            22,
            max(len(column), *[len(_fmt(row.get(column))) for row in rows]),
        )
        for column in columns
    }
    output = [
        " | ".join(column.ljust(widths[column]) for column in columns),
        "-+-".join("-" * widths[column] for column in columns),
    ]
    for row in rows:
        output.append(
            " | ".join(
                _fmt(row.get(column))[: widths[column]].ljust(widths[column])
                for column in columns
            )
        )
    return output


def _fmt(value):
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float):
        return f"{value:.6f}"
    return str(value)
